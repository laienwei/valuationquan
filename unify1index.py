#!/usr/bin/python3
# coding: utf-8

# In[28]:
import openpyxl
import xlrd
from openpyxl.styles import Font,Alignment
datelai='2023-09-28'
fontText1 = Font(name='Tahoma', size=10, color='ff0000')
AlignmentText1 = Alignment(horizontal='center')
wb0 = openpyxl.load_workbook(r'./lai/valuationquan/szseinnovation100index.xlsx')
##选择要操作的工作表， 返回工作表对象
sheet0 = wb0['szse_innovation_100']
xx=sheet0.max_row 
index=sheet0.cell(row=xx,column=5).value/1000
filedir3=['szseinnovation100indexmodel1.xlsx','szseinnovation100indexmodel2.xlsx','szseinnovation100indexmodel4.xlsx']
dirba1="./lai/valuationquan/szseinnovation100index"
filename3=dirba1+"/"+filedir3[0]
data=xlrd.open_workbook(filename3)#openpyxl模块读不出数据，只能读出单元格公式
sheet1_data=data.sheet_by_name('model1')
wb = openpyxl.load_workbook(filename3)
sheet = wb['model1']
p=sheet.max_row
print(p)
print(sheet1_data.cell_value(p-1,4))
aa=sheet1_data.cell_value(p-1,4)
print(type(aa))
print(sheet.cell(row=p,column=5).value)
print(sheet1_data.cell_value(p-1,6))
bb=sheet1_data.cell_value(p-1,6)
print(type(bb))
print(sheet.cell(row=p,column=7).value)
sheet.cell(row=p+1, column=1).value=datelai
sheet.cell(row=p+1, column=1).alignment=AlignmentText1
sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
sheet.cell(row=p+1, column=1).font=fontText1
sheet.cell(row=p+1, column=3).value=2000
sheet.cell(row=p+1, column=3).font=fontText1
sheet.cell(row=p+1, column=3).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=2).value=index
sheet.cell(row=p+1, column=2).font=fontText1
sheet.cell(row=p+1, column=2).number_format = '0.00000'
sheet.cell(row=p+1, column=4).value=2000/index
sheet.cell(row=p+1, column=4).font=fontText1
sheet.cell(row=p+1, column=4).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=5).value=2000/index+float(aa)
sheet.cell(row=p+1, column=5).font=fontText1
sheet.cell(row=p+1, column=5).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=6).value=(2000/index+float(aa))*index
sheet.cell(row=p+1, column=6).font=fontText1
sheet.cell(row=p+1, column=6).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=7).value=float(bb)+2000
sheet.cell(row=p+1, column=7).font=fontText1
sheet.cell(row=p+1, column=7).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=8).value=(2000/index+float(aa))*index
sheet.cell(row=p+1, column=8).font=fontText1
sheet.cell(row=p+1, column=8).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=(2000/index+float(aa))*index-float(bb)-2000
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_ '
wb.save(filename3)


# In[38]:
filename4=dirba1+"/"+filedir3[1]
data=xlrd.open_workbook(filename4)#openpyxl模块读不出数据，只能读出单元格公式
sheet1_data=data.sheet_by_name('model2(1)')
filename3=dirba1+"/"+"szseinnovation100indexmodel1prac.xlsx"
data1=xlrd.open_workbook(filename3)#openpyxl模块读不出数据，只能读出单元格公式
sheet2_data=data1.sheet_by_name('myPEPB')
nrow=sheet2_data.nrows

wb = openpyxl.load_workbook(filename4)
sheet = wb['model2(1)']
p=sheet.max_row
print(p)
print(sheet1_data.cell_value(p-1,6))
aa=sheet1_data.cell_value(p-1,6)
print(sheet1_data.cell_value(p-1,8))
bb=sheet1_data.cell_value(p-1,8)
print(sheet1_data.cell_value(p-1,11))
cc=sheet1_data.cell_value(p-1,11)
print(type(sheet2_data.cell_value(nrow-1,2)))
print(sheet2_data.cell_value(nrow-1,2))
print(type(sheet2_data.cell_value(nrow-1,3)))
print(float(sheet2_data.cell_value(nrow-1,3)))
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2)) * 3950
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2)) * -3950
    biaozhi=0
sheet.cell(row=p+1, column=1).value=datelai
sheet.cell(row=p+1, column=1).alignment=AlignmentText1
sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
sheet.cell(row=p+1, column=1).font=fontText1
sheet.cell(row=p+1, column=2).value=index
sheet.cell(row=p+1, column=2).font=fontText1
sheet.cell(row=p+1, column=2).number_format = '0.00000'
sheet.cell(row=p+1, column=3).value=sheet2_data.cell_value(nrow-1,2)#PE值
sheet.cell(row=p+1, column=3).font=fontText1
sheet.cell(row=p+1, column=3).number_format = 'General'
sheet.cell(row=p+1, column=4).value=sheet2_data.cell_value(nrow-1,3)#PE均值
sheet.cell(row=p+1, column=4).font=fontText1
sheet.cell(row=p+1, column=4).number_format = 'General'
sheet.cell(row=p+1, column=5).value=amount
sheet.cell(row=p+1, column=5).font=fontText1
sheet.cell(row=p+1, column=5).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=6).value=amount/index
sheet.cell(row=p+1, column=6).font=fontText1
sheet.cell(row=p+1, column=6).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=7).value=amount/index+float(aa)
sheet.cell(row=p+1, column=7).font=fontText1
sheet.cell(row=p+1, column=7).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=8).value=(amount/index+float(aa))*index
sheet.cell(row=p+1, column=8).font=fontText1
sheet.cell(row=p+1, column=8).number_format = '0.00_);-0.00'
if(biaozhi==1):
    sheet.cell(row=p+1, column=9).value=float(bb)+amount
    sheet.cell(row=p+1, column=12).value=float(cc)
else:
    sheet.cell(row=p+1, column=9).value=float(bb)
    sheet.cell(row=p+1, column=12).value=float(cc)-amount
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=8).value+sheet.cell(row=p+1, column=12).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=10).value-sheet.cell(row=p+1, column=9).value
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_ '
wb.save(filename4)


filename5=dirba1+"/"+filedir3[2]
data=xlrd.open_workbook(filename5)#openpyxl模块读不出数据，只能读出单元格公式
sheet1_data=data.sheet_by_name('model4(1)')
#filename3=dirba1+"/"+"szseinnovation100indexmodel1prac.xlsx"
data1=xlrd.open_workbook(filename3)#openpyxl模块读不出数据，只能读出单元格公式
sheet2_data=data1.sheet_by_name('myPEPB')
nrow=sheet2_data.nrows

wb = openpyxl.load_workbook(filename5)
sheet = wb['model4(1)']
p=sheet.max_row
print(p)
print(sheet1_data.cell_value(p-1,6))
aa=sheet1_data.cell_value(p-1,6)
print(sheet1_data.cell_value(p-1,8))
bb=sheet1_data.cell_value(p-1,8)
print(sheet1_data.cell_value(p-1,11))
cc=sheet1_data.cell_value(p-1,11)
print(type(sheet2_data.cell_value(nrow-1,2)))
print(sheet2_data.cell_value(nrow-1,2))
print(type(sheet2_data.cell_value(nrow-1,3)))
print(float(sheet2_data.cell_value(nrow-1,3)))
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))**2 * 3950
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))**2 * -3950
    biaozhi=0
sheet.cell(row=p+1, column=1).value=datelai
sheet.cell(row=p+1, column=1).alignment=AlignmentText1
sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
sheet.cell(row=p+1, column=1).font=fontText1
sheet.cell(row=p+1, column=2).value=index
sheet.cell(row=p+1, column=2).font=fontText1
sheet.cell(row=p+1, column=2).number_format = '0.00000'
sheet.cell(row=p+1, column=3).value=sheet2_data.cell_value(nrow-1,2)#PE值
sheet.cell(row=p+1, column=3).font=fontText1
sheet.cell(row=p+1, column=3).number_format = 'General'
sheet.cell(row=p+1, column=4).value=sheet2_data.cell_value(nrow-1,3)#PE均值
sheet.cell(row=p+1, column=4).font=fontText1
sheet.cell(row=p+1, column=4).number_format = 'General'
sheet.cell(row=p+1, column=5).value=amount
sheet.cell(row=p+1, column=5).font=fontText1
sheet.cell(row=p+1, column=5).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=6).value=amount/index
sheet.cell(row=p+1, column=6).font=fontText1
sheet.cell(row=p+1, column=6).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=7).value=amount/index+float(aa)
sheet.cell(row=p+1, column=7).font=fontText1
sheet.cell(row=p+1, column=7).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=8).value=(amount/index+float(aa))*index
sheet.cell(row=p+1, column=8).font=fontText1
sheet.cell(row=p+1, column=8).number_format = '0.00_);-0.00'
if(biaozhi==1):
    sheet.cell(row=p+1, column=9).value=float(bb)+amount
    sheet.cell(row=p+1, column=12).value=float(cc)
else:
    sheet.cell(row=p+1, column=9).value=float(bb)
    sheet.cell(row=p+1, column=12).value=float(cc)-amount
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=8).value+sheet.cell(row=p+1, column=12).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=10).value-sheet.cell(row=p+1, column=9).value
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_ '
wb.save(filename5)


# In[ ]:




