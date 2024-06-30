#!/usr/bin/env python
# coding: utf-8

# In[ ]:


#!/usr/bin/python3
# coding: utf-8

# In[15]:


import openpyxl
import xlrd
from openpyxl.styles import Font,Alignment
import publai1
pub=publai1.publai1()
datelai=pub.date
a=[]
b=[]
c=[]
d=[]
fontText1 = Font(name='Tahoma', size=10, color='ff0000')
fontText2 = Font(name='Tahoma', size=10)
AlignmentText1 = Alignment(horizontal='center')
wb0 = xlrd.open_workbook(r'./lai/valuationquan/RTcsiAIindexLOF.xlsx')
##选择要操作的工作表， 返回工作表对象
sheet0 = wb0.sheet_by_name('RT_csiAI_index')
xx=sheet0.nrows 
price=sheet0.cell_value(xx-1,4)
pricemean=sheet0.cell_value(xx-1,8)
vol=sheet0.cell_value(xx-1,5)
volmean=sheet0.cell_value(xx-1,13)
high=max(sheet0.cell_value(xx-1,2),sheet0.cell_value(xx-2,11))
low=min(sheet0.cell_value(xx-1,3),sheet0.cell_value(xx-2,12))
filedir3=['csiAILOFmodel1mean.xlsx','csiAILOFmodel2mean.xlsx','csiAILOFmodel4mean.xlsx']
dirba1="./lai/valuationquan/AILOF"
filename3=dirba1+"/"+filedir3[0]
data=xlrd.open_workbook(filename3)#openpyxl模块读不出数据，只能读出单元格公式
wb = openpyxl.load_workbook(filename3)


# In[16]:


for sheet in wb:
    print(sheet.title)
    p=sheet.max_row
    print(p)
    sheet.cell(row=p+1, column=1).value=datelai
    sheet.cell(row=p+1, column=1).alignment=AlignmentText1
    sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
    sheet.cell(row=p+1, column=1).font=fontText1
    sheet.cell(row=p+1, column=2).value=price
    sheet.cell(row=p+1, column=2).font=fontText1
    sheet.cell(row=p+1, column=2).number_format = '0.00000'
    a.append(sheet.title)
    b.append(data.sheet_by_name(sheet.title).cell_value(p-2,4))
    c.append(data.sheet_by_name(sheet.title).cell_value(p-2,6))  
mydict1=dict(zip(a,b))
mydict2=dict(zip(a,c))
###################################################################
sheet = wb['model1']
sheet.cell(row=p+1, column=3).value=2000
sheet.cell(row=p+1, column=4).value=2000/price
sheet.cell(row=p+1, column=5).value=2000/price+float(mydict1['model1'])
sheet.cell(row=p+1, column=6).value=(2000/price+float(mydict1['model1']))*price
sheet.cell(row=p+1, column=7).value=float(mydict2['model1'])+2000
sheet.cell(row=p+1, column=8).value=(2000/price+float(mydict1['model1']))*price
sheet.cell(row=p+1, column=9).value=(2000/price+float(mydict1['model1']))*price-float(mydict2['model1'])-2000
for i in range(6):
    sheet.cell(row=p+1, column=i+3).font=fontText1
    sheet.cell(row=p+1, column=i+3).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_ '
###################################################################
sheet = wb['model1&RSI']
tt=price-float(data.sheet_by_name(sheet.title).cell_value(p-2,1))
sheet.cell(row=p+1, column=10).value=max(tt,0)
sheet.cell(row=p+1, column=11).value=(max(tt,0)+float(data.sheet_by_name(sheet.title).cell_value(p-2,10))*5)/6
sheet.cell(row=p+1, column=12).value=abs(tt)
sheet.cell(row=p+1, column=13).value=(abs(tt)+float(data.sheet_by_name(sheet.title).cell_value(p-2,12))*5)/6
sma1=(max(tt,0)+float(data.sheet_by_name(sheet.title).cell_value(p-2,10))*5)/6
sma2=(abs(tt)+float(data.sheet_by_name(sheet.title).cell_value(p-2,12))*5)/6
sheet.cell(row=p+1, column=14).value=sma1*100/sma2
for i in range(5):
    sheet.cell(row=p+1, column=i+10).font=fontText2
    sheet.cell(row=p+1, column=i+10).number_format = 'General'
if((sma1*100/sma2)<20):
    sheet.cell(row=p+1, column=3).value=3000
    sheet.cell(row=p+1, column=4).value=3000/price
    sheet.cell(row=p+1, column=5).value=3000/price+float(mydict1['model1&RSI'])
    sheet.cell(row=p+1, column=6).value=(3000/price+float(mydict1['model1&RSI']))*price
    sheet.cell(row=p+1, column=7).value=3000+float(mydict2['model1&RSI'])
    sheet.cell(row=p+1, column=8).value=(3000/price+float(mydict1['model1&RSI']))*price
    sheet.cell(row=p+1, column=9).value=(3000/price+float(mydict1['model1&RSI']))*price-float(mydict2['model1&RSI'])-3000
else:
    sheet.cell(row=p+1, column=3).value=2000
    sheet.cell(row=p+1, column=4).value=2000/price
    sheet.cell(row=p+1, column=5).value=2000/price+float(mydict1['model1&RSI'])
    sheet.cell(row=p+1, column=6).value=(2000/price+float(mydict1['model1&RSI']))*price
    sheet.cell(row=p+1, column=7).value=2000+float(mydict2['model1&RSI'])
    sheet.cell(row=p+1, column=8).value=(2000/price+float(mydict1['model1&RSI']))*price
    sheet.cell(row=p+1, column=9).value=(2000/price+float(mydict1['model1&RSI']))*price-float(mydict2['model1&RSI'])-2000
for i in range(6):
    sheet.cell(row=p+1, column=i+3).font=fontText1
    sheet.cell(row=p+1, column=i+3).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_ '

###########################################################################
sheet = wb['model1&KDJ']
sheet.cell(row=p+1, column=10).value=high
sheet.cell(row=p+1, column=11).value=low
maxlai=high
minlai=low
t=0
while t<8:
    maxlai=max(maxlai,data.sheet_by_name(sheet.title).cell_value(p-2-t,9))
    minlai=min(minlai,data.sheet_by_name(sheet.title).cell_value(p-2-t,10))
    t=t+1
print("laienwei;")
print(maxlai)
print(minlai)
sheet.cell(row=p+1, column=12).value=maxlai
sheet.cell(row=p+1, column=13).value=minlai
sheet.cell(row=p+1, column=14).value=(price-minlai)*100/(maxlai-minlai)
K=(2*data.sheet_by_name(sheet.title).cell_value(p-2,14)+(price-minlai)*100/(maxlai-minlai))/3
D=(2*data.sheet_by_name(sheet.title).cell_value(p-2,15)+K)/3
J=3*K-2*D
sheet.cell(row=p+1, column=15).value=K
sheet.cell(row=p+1, column=16).value=D
sheet.cell(row=p+1, column=17).value=J
for i in range(8):
    sheet.cell(row=p+1, column=10+i).font=fontText2
    sheet.cell(row=p+1, column=10+i).number_format = 'General'
if(J<0):
    sheet.cell(row=p+1, column=3).value=3000
    sheet.cell(row=p+1, column=4).value=3000/price
    sheet.cell(row=p+1, column=5).value=3000/price+float(mydict1['model1&KDJ'])
    sheet.cell(row=p+1, column=6).value=(3000/price+float(mydict1['model1&KDJ']))*price
    sheet.cell(row=p+1, column=7).value=3000+float(mydict2['model1&KDJ'])  
    sheet.cell(row=p+1, column=8).value=(3000/price+float(mydict1['model1&KDJ']))*price
    sheet.cell(row=p+1, column=9).value=(3000/price+float(mydict1['model1&KDJ']))*price-float(mydict2['model1&KDJ'])-3000
else:
    sheet.cell(row=p+1, column=3).value=2000
    sheet.cell(row=p+1, column=4).value=2000/price
    sheet.cell(row=p+1, column=5).value=2000/price+float(mydict1['model1&KDJ'])
    sheet.cell(row=p+1, column=6).value=(2000/price+float(mydict1['model1&KDJ']))*price
    sheet.cell(row=p+1, column=7).value=2000+float(mydict2['model1&KDJ'])
    sheet.cell(row=p+1, column=8).value=(2000/price+float(mydict1['model1&KDJ']))*price
    sheet.cell(row=p+1, column=9).value=(2000/price+float(mydict1['model1&KDJ']))*price-float(mydict2['model1&KDJ'])-2000
for i in range(6):
    sheet.cell(row=p+1, column=i+3).font=fontText1
    sheet.cell(row=p+1, column=i+3).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_ '
wb.save(filename3)
###########################################################################



# In[27]:

a=[]
b=[]
c=[]
d=[]
filename4=dirba1+"/"+filedir3[1]
data=xlrd.open_workbook(filename4)#openpyxl模块读不出数据，只能读出单元格公式
wb = openpyxl.load_workbook(filename4)
sheet = wb['model2(1)']
p=sheet.max_row
print(p)
for sheet in wb:
    print(sheet.title)
    p=sheet.max_row
    sheet.cell(row=p+1, column=1).value=datelai
    sheet.cell(row=p+1, column=1).alignment=AlignmentText1
    sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
    sheet.cell(row=p+1, column=1).font=fontText1
    sheet.cell(row=p+1, column=2).value=price
    sheet.cell(row=p+1, column=2).font=fontText1
    sheet.cell(row=p+1, column=2).number_format = '0.00000'
    sheet.cell(row=p+1, column=3).value=pricemean
    sheet.cell(row=p+1, column=3).font=fontText1
    sheet.cell(row=p+1, column=3).number_format = 'General'
    if((sheet.title!='model2(1)vol')and(sheet.title!='model2(2)vol')):
        a.append(sheet.title)
        b.append(data.sheet_by_name(sheet.title).cell_value(p-2,5))
        c.append(data.sheet_by_name(sheet.title).cell_value(p-2,7))
        d.append(data.sheet_by_name(sheet.title).cell_value(p-2,10))
    else:
        a.append(sheet.title)
        b.append(data.sheet_by_name(sheet.title).cell_value(p-2,7))
        c.append(data.sheet_by_name(sheet.title).cell_value(p-2,9))
        d.append(data.sheet_by_name(sheet.title).cell_value(p-2,12))   
mydict1=dict(zip(a,b))
mydict2=dict(zip(a,c))
mydict3=dict(zip(a,d))
print(mydict1)
print(mydict2)
print(mydict3)
#########################################################################
sheet = wb['model2(1)']
aa=float(mydict1['model2(1)'])
bb=float(mydict2['model2(1)'])
cc=float(mydict3['model2(1)'])
if (price<=pricemean):
    amount=(pricemean-price) * 1550
    biaozhi=1
else:
    amount=(price-pricemean) * -1550
    biaozhi=0
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=5).value=amount/price
sheet.cell(row=p+1, column=6).value=amount/price+float(aa)
sheet.cell(row=p+1, column=7).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
############################################################################
sheet = wb['model2(2)']
aa=float(mydict1['model2(2)'])
bb=float(mydict2['model2(2)'])
cc=float(mydict3['model2(2)'])
if (price<=pricemean):
    amount=(pricemean-price)**2 * 1550
    biaozhi=1
else:
    amount=(price-pricemean)**2 * -1550
    biaozhi=0
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=5).value=amount/price
sheet.cell(row=p+1, column=6).value=amount/price+float(aa)
sheet.cell(row=p+1, column=7).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
###########################################################################
sheet = wb['model2(1)vol']
aa=float(mydict1['model2(1)vol'])
bb=float(mydict2['model2(1)vol'])
cc=float(mydict3['model2(1)vol'])
if (price<=pricemean):
    amount=(pricemean-price) * (vol/volmean) * 1550
    biaozhi=1
else:
    amount=(price-pricemean) * (vol/volmean) * -1550
    biaozhi=0
sheet.cell(row=p+1, column=4).value=vol
sheet.cell(row=p+1, column=5).value=volmean
sheet.cell(row=p+1, column=6).value=amount
sheet.cell(row=p+1, column=7).value=amount/price
sheet.cell(row=p+1, column=8).value=amount/price+float(aa)
sheet.cell(row=p+1, column=9).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=10).value=float(bb)+amount
    sheet.cell(row=p+1, column=13).value=float(cc)
else:
    sheet.cell(row=p+1, column=10).value=float(bb)
    sheet.cell(row=p+1, column=13).value=float(cc)-amount
sheet.cell(row=p+1, column=13).font=fontText1
sheet.cell(row=p+1, column=13).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=9).value+sheet.cell(row=p+1, column=13).value
sheet.cell(row=p+1, column=12).value=sheet.cell(row=p+1, column=11).value-sheet.cell(row=p+1, column=10).value
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_ '
for i in range(8):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
############################################################################
sheet = wb['model2(2)vol']
aa=float(mydict1['model2(2)vol'])
bb=float(mydict2['model2(2)vol'])
cc=float(mydict3['model2(2)vol'])
if (price<=pricemean):
    amount=(pricemean-price) **2 * (vol/volmean) * 1550
    biaozhi=1
else:
    amount=(price-pricemean) **2 * (vol/volmean) * -1550
    biaozhi=0
sheet.cell(row=p+1, column=4).value=vol
sheet.cell(row=p+1, column=5).value=volmean
sheet.cell(row=p+1, column=6).value=amount
sheet.cell(row=p+1, column=7).value=amount/price
sheet.cell(row=p+1, column=8).value=amount/price+float(aa)
sheet.cell(row=p+1, column=9).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=10).value=float(bb)+amount
    sheet.cell(row=p+1, column=13).value=float(cc)
else:
    sheet.cell(row=p+1, column=10).value=float(bb)
    sheet.cell(row=p+1, column=13).value=float(cc)-amount
sheet.cell(row=p+1, column=13).font=fontText1
sheet.cell(row=p+1, column=13).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=9).value+sheet.cell(row=p+1, column=13).value
sheet.cell(row=p+1, column=12).value=sheet.cell(row=p+1, column=11).value-sheet.cell(row=p+1, column=10).value
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_ '
for i in range(8):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
#####################################################################
sheet = wb['model2(1)&RSI']
aa=float(mydict1['model2(1)&RSI'])
bb=float(mydict2['model2(1)&RSI'])
cc=float(mydict3['model2(1)&RSI'])
datax=xlrd.open_workbook(filename3)#openpyxl模块读不出数据，只能读出单元格公式
sheetx=datax.sheet_by_name('model1&RSI')
nrowx=sheetx.nrows
sheet.cell(row=p+1, column=12).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,10)
sheet.cell(row=p+1, column=13).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,12)
sheet.cell(row=p+1, column=14).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
RSIpre=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
if(float(RSIpre)<20):
    sign=2
elif((float(RSIpre)>20)and(RSIpre<25)):
    sign=1
elif((float(RSIpre)>25)and(RSIpre<50)):
    sign=0.95
elif((float(RSIpre)>50)and(RSIpre<80)):
    sign=0.2
else:
    sign=1
sheet.cell(row=p+1, column=15).value=sign
sheet.cell(row=p+1, column=15).font=fontText2
sheet.cell(row=p+1, column=15).number_format = 'General'
if (price<=pricemean):
    amount=(pricemean-price) * 1550 * sign
    biaozhi=1
else:
    amount=(price-pricemean) * -1550 * sign
    biaozhi=0
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=5).value=amount/price
sheet.cell(row=p+1, column=6).value=amount/price+float(aa)
sheet.cell(row=p+1, column=7).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
for i in range(4):
    sheet.cell(row=p+1, column=i+12).font=fontText1
    sheet.cell(row=p+1, column=i+12).number_format = '0.000000_);(0.000000)' 
#####################################################################
sheet = wb['model2(2)&RSI']
aa=float(mydict1['model2(2)&RSI'])
bb=float(mydict2['model2(2)&RSI'])
cc=float(mydict3['model2(2)&RSI'])
#datax=xlrd.open_workbook(filename3)#openpyxl模块读不出数据，只能读出单元格公式
#sheetx=datax.sheet_by_name('model1&RSI')
#nrowx=sheetx.nrows
sheet.cell(row=p+1, column=12).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,10)
sheet.cell(row=p+1, column=13).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,12)
sheet.cell(row=p+1, column=14).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
RSIpre=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
if(float(RSIpre)<20):
    sign=2
elif((float(RSIpre)>20)and(RSIpre<25)):
    sign=1
elif((float(RSIpre)>25)and(RSIpre<50)):
    sign=0.95
elif((float(RSIpre)>50)and(RSIpre<80)):
    sign=0.2
else:
    sign=1
sheet.cell(row=p+1, column=15).value=sign
sheet.cell(row=p+1, column=15).font=fontText2
sheet.cell(row=p+1, column=15).number_format = 'General'
if (price<=pricemean):
    amount=(pricemean-price)**2 * 1550 * sign
    biaozhi=1
else:
    amount=(price-pricemean)**2 * -1550 * sign
    biaozhi=0
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=5).value=amount/price
sheet.cell(row=p+1, column=6).value=amount/price+float(aa)
sheet.cell(row=p+1, column=7).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
for i in range(4):
    sheet.cell(row=p+1, column=i+12).font=fontText1
    sheet.cell(row=p+1, column=i+12).number_format = '0.000000_);(0.000000)'
############################################################################
sheet = wb['model2(1)&KDJ']
aa=float(mydict1['model2(1)&KDJ'])
bb=float(mydict2['model2(1)&KDJ'])
cc=float(mydict3['model2(1)&KDJ'])
sheet.cell(row=p+1, column=12).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,13)
sheet.cell(row=p+1, column=13).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,14)
sheet.cell(row=p+1, column=14).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,15)
sheet.cell(row=p+1, column=15).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,16)
KDJ=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,16)
if((KDJ<-10)or(KDJ>110)):
    sign=1
elif((KDJ<55)and(KDJ>45)):
    sign=0.2
else:
    sign=0.95
sheet.cell(row=p+1, column=16).value=sign
if (price<=pricemean):
    amount=(pricemean-price) * 1550 * sign
    biaozhi=1
else:
    amount=(price-pricemean) * -1550 * sign
    biaozhi=0
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=5).value=amount/price
sheet.cell(row=p+1, column=6).value=amount/price+float(aa)
sheet.cell(row=p+1, column=7).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
for i in range(4):
    sheet.cell(row=p+1, column=i+12).font=fontText1
    sheet.cell(row=p+1, column=i+12).number_format = '0.00_);-0.00'   
sheet.cell(row=p+1, column=15).font=fontText1
sheet.cell(row=p+1, column=15).number_format = '0.00_ '  
sheet.cell(row=p+1, column=16).font=fontText2
sheet.cell(row=p+1, column=16).number_format = 'General'
############################################################################
sheet = wb['model2(2)&KDJ']
aa=float(mydict1['model2(2)&KDJ'])
bb=float(mydict2['model2(2)&KDJ'])
cc=float(mydict3['model2(2)&KDJ'])
sheet.cell(row=p+1, column=12).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,13)
sheet.cell(row=p+1, column=13).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,14)
sheet.cell(row=p+1, column=14).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,15)
sheet.cell(row=p+1, column=15).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,16)
KDJ=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,16)
if((KDJ<-10)or(KDJ>110)):
    sign=1
elif((KDJ<55)and(KDJ>45)):
    sign=0.2
else:
    sign=0.95
sheet.cell(row=p+1, column=16).value=sign
if (price<=pricemean):
    amount=(pricemean-price)**2 * 1550 * sign
    biaozhi=1
else:
    amount=(price-pricemean)**2 * -1550 * sign
    biaozhi=0
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=5).value=amount/price
sheet.cell(row=p+1, column=6).value=amount/price+float(aa)
sheet.cell(row=p+1, column=7).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
for i in range(4):
    sheet.cell(row=p+1, column=i+12).font=fontText1
    sheet.cell(row=p+1, column=i+12).number_format = '0.00_);-0.00'   
sheet.cell(row=p+1, column=15).font=fontText1
sheet.cell(row=p+1, column=15).number_format = '0.00_ '  
sheet.cell(row=p+1, column=16).font=fontText2
sheet.cell(row=p+1, column=16).number_format = 'General'

wb.save(filename4)

# In[27]:


# In[ ]:


a=[]
b=[]
c=[]
d=[]
filename5=dirba1+"/"+filedir3[2]
data=xlrd.open_workbook(filename5)#openpyxl模块读不出数据，只能读出单元格公式
wb = openpyxl.load_workbook(filename5)
sheet = wb['model4(1)']
p=sheet.max_row
print(p)
for sheet in wb:
    print(sheet.title)
    p=sheet.max_row
    sheet.cell(row=p+1, column=1).value=datelai
    sheet.cell(row=p+1, column=1).alignment=AlignmentText1
    sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
    sheet.cell(row=p+1, column=1).font=fontText1
    sheet.cell(row=p+1, column=2).value=price
    sheet.cell(row=p+1, column=2).font=fontText1
    sheet.cell(row=p+1, column=2).number_format = '0.00000'
    sheet.cell(row=p+1, column=3).value=pricemean
    sheet.cell(row=p+1, column=3).font=fontText1
    sheet.cell(row=p+1, column=3).number_format = 'General'
    if((sheet.title!='model4(1)vol')and(sheet.title!='model4(3)vol')):
        a.append(sheet.title)
        b.append(data.sheet_by_name(sheet.title).cell_value(p-2,5))
        c.append(data.sheet_by_name(sheet.title).cell_value(p-2,7))
        d.append(data.sheet_by_name(sheet.title).cell_value(p-2,10))
    else:
        a.append(sheet.title)
        b.append(data.sheet_by_name(sheet.title).cell_value(p-2,7))
        c.append(data.sheet_by_name(sheet.title).cell_value(p-2,9))
        d.append(data.sheet_by_name(sheet.title).cell_value(p-2,12))   
mydict1=dict(zip(a,b))
mydict2=dict(zip(a,c))
mydict3=dict(zip(a,d))
print(mydict1)
print(mydict2)
print(mydict3)
#########################################################################
sheet = wb['model4(1)']
aa=float(mydict1['model4(1)'])
bb=float(mydict2['model4(1)'])
cc=float(mydict3['model4(1)'])
if (price<=pricemean):
    amount=(pricemean-price)**2 * 1550
    biaozhi=1
else:
    amount=(price-pricemean)**2 * -1550
    biaozhi=0
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=5).value=amount/price
sheet.cell(row=p+1, column=6).value=amount/price+float(aa)
sheet.cell(row=p+1, column=7).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
############################################################################
sheet = wb['model4(3)']
aa=float(mydict1['model4(3)'])
bb=float(mydict2['model4(3)'])
cc=float(mydict3['model4(3)'])
if (price<=pricemean):
    amount=(pricemean-price)**3 * 1550
    biaozhi=1
else:
    amount=(price-pricemean)**3 * -1550
    biaozhi=0
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=5).value=amount/price
sheet.cell(row=p+1, column=6).value=amount/price+float(aa)
sheet.cell(row=p+1, column=7).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
###########################################################################
sheet = wb['model4(1)vol']
aa=float(mydict1['model4(1)vol'])
bb=float(mydict2['model4(1)vol'])
cc=float(mydict3['model4(1)vol'])
if (price<=pricemean):
    amount=(pricemean-price) * (vol/volmean)**2 * 1550
    biaozhi=1
else:
    amount=(price-pricemean) * (vol/volmean)**2 * -1550
    biaozhi=0
sheet.cell(row=p+1, column=4).value=vol
sheet.cell(row=p+1, column=5).value=volmean
sheet.cell(row=p+1, column=6).value=amount
sheet.cell(row=p+1, column=7).value=amount/price
sheet.cell(row=p+1, column=8).value=amount/price+float(aa)
sheet.cell(row=p+1, column=9).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=10).value=float(bb)+amount
    sheet.cell(row=p+1, column=13).value=float(cc)
else:
    sheet.cell(row=p+1, column=10).value=float(bb)
    sheet.cell(row=p+1, column=13).value=float(cc)-amount
sheet.cell(row=p+1, column=13).font=fontText1
sheet.cell(row=p+1, column=13).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=9).value+sheet.cell(row=p+1, column=13).value
sheet.cell(row=p+1, column=12).value=sheet.cell(row=p+1, column=11).value-sheet.cell(row=p+1, column=10).value
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_ '
for i in range(8):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
############################################################################
sheet = wb['model4(3)vol']
aa=float(mydict1['model4(3)vol'])
bb=float(mydict2['model4(3)vol'])
cc=float(mydict3['model4(3)vol'])
if (price<=pricemean):
    amount=(pricemean-price) **3 * (vol/volmean) * 1550
    biaozhi=1
else:
    amount=(price-pricemean) **3 * (vol/volmean) * -1550
    biaozhi=0
sheet.cell(row=p+1, column=4).value=vol
sheet.cell(row=p+1, column=5).value=volmean
sheet.cell(row=p+1, column=6).value=amount
sheet.cell(row=p+1, column=7).value=amount/price
sheet.cell(row=p+1, column=8).value=amount/price+float(aa)
sheet.cell(row=p+1, column=9).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=10).value=float(bb)+amount
    sheet.cell(row=p+1, column=13).value=float(cc)
else:
    sheet.cell(row=p+1, column=10).value=float(bb)
    sheet.cell(row=p+1, column=13).value=float(cc)-amount
sheet.cell(row=p+1, column=13).font=fontText1
sheet.cell(row=p+1, column=13).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=9).value+sheet.cell(row=p+1, column=13).value
sheet.cell(row=p+1, column=12).value=sheet.cell(row=p+1, column=11).value-sheet.cell(row=p+1, column=10).value
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_ '
for i in range(8):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
#####################################################################
sheet = wb['model4(1)&RSI']
aa=float(mydict1['model4(1)&RSI'])
bb=float(mydict2['model4(1)&RSI'])
cc=float(mydict3['model4(1)&RSI'])
datax=xlrd.open_workbook(filename3)#openpyxl模块读不出数据，只能读出单元格公式
sheetx=datax.sheet_by_name('model1&RSI')
nrowx=sheetx.nrows
sheet.cell(row=p+1, column=12).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,10)
sheet.cell(row=p+1, column=13).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,12)
sheet.cell(row=p+1, column=14).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
RSIpre=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
if(float(RSIpre)<20):
    sign=2
elif((float(RSIpre)>20)and(RSIpre<25)):
    sign=1
elif((float(RSIpre)>25)and(RSIpre<50)):
    sign=0.95
elif((float(RSIpre)>50)and(RSIpre<80)):
    sign=0.2
else:
    sign=1
sheet.cell(row=p+1, column=15).value=sign
sheet.cell(row=p+1, column=15).font=fontText2
sheet.cell(row=p+1, column=15).number_format = 'General'
if (price<=pricemean):
    amount=(pricemean-price)**2 * 1550 * sign
    biaozhi=1
else:
    amount=(price-pricemean)**2 * -1550 * sign
    biaozhi=0
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=5).value=amount/price
sheet.cell(row=p+1, column=6).value=amount/price+float(aa)
sheet.cell(row=p+1, column=7).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
for i in range(4):
    sheet.cell(row=p+1, column=i+12).font=fontText1
    sheet.cell(row=p+1, column=i+12).number_format = '0.000000_);(0.000000)' 
#####################################################################
sheet = wb['model4(3)&RSI']
aa=float(mydict1['model4(3)&RSI'])
bb=float(mydict2['model4(3)&RSI'])
cc=float(mydict3['model4(3)&RSI'])
#datax=xlrd.open_workbook(filename3)#openpyxl模块读不出数据，只能读出单元格公式
#sheetx=datax.sheet_by_name('model1&RSI')
#nrowx=sheetx.nrows
sheet.cell(row=p+1, column=12).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,10)
sheet.cell(row=p+1, column=13).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,12)
sheet.cell(row=p+1, column=14).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
RSIpre=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
if(float(RSIpre)<20):
    sign=2
elif((float(RSIpre)>20)and(RSIpre<25)):
    sign=1
elif((float(RSIpre)>25)and(RSIpre<50)):
    sign=0.95
elif((float(RSIpre)>50)and(RSIpre<80)):
    sign=0.2
else:
    sign=1
sheet.cell(row=p+1, column=15).value=sign
sheet.cell(row=p+1, column=15).font=fontText2
sheet.cell(row=p+1, column=15).number_format = 'General'
if (price<=pricemean):
    amount=(pricemean-price)**3 * 1550 * sign
    biaozhi=1
else:
    amount=(price-pricemean)**3 * -1550 * sign
    biaozhi=0
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=5).value=amount/price
sheet.cell(row=p+1, column=6).value=amount/price+float(aa)
sheet.cell(row=p+1, column=7).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
for i in range(4):
    sheet.cell(row=p+1, column=i+12).font=fontText1
    sheet.cell(row=p+1, column=i+12).number_format = '0.000000_);(0.000000)'
############################################################################
sheet = wb['model4(1)&KDJ']
aa=float(mydict1['model4(1)&KDJ'])
bb=float(mydict2['model4(1)&KDJ'])
cc=float(mydict3['model4(1)&KDJ'])
sheet.cell(row=p+1, column=12).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,13)
sheet.cell(row=p+1, column=13).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,14)
sheet.cell(row=p+1, column=14).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,15)
sheet.cell(row=p+1, column=15).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,16)
KDJ=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,16)
if((KDJ<-10)or(KDJ>110)):
    sign=1
elif((KDJ<55)and(KDJ>45)):
    sign=0.2
else:
    sign=0.95
sheet.cell(row=p+1, column=16).value=sign
if (price<=pricemean):
    amount=(pricemean-price)**2 * 1550 * sign
    biaozhi=1
else:
    amount=(price-pricemean)**2 * -1550 * sign
    biaozhi=0
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=5).value=amount/price
sheet.cell(row=p+1, column=6).value=amount/price+float(aa)
sheet.cell(row=p+1, column=7).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
for i in range(4):
    sheet.cell(row=p+1, column=i+12).font=fontText1
    sheet.cell(row=p+1, column=i+12).number_format = '0.00_);-0.00'   
sheet.cell(row=p+1, column=15).font=fontText1
sheet.cell(row=p+1, column=15).number_format = '0.00_ '  
sheet.cell(row=p+1, column=16).font=fontText2
sheet.cell(row=p+1, column=16).number_format = 'General'
############################################################################
sheet = wb['model4(3)&KDJ']
aa=float(mydict1['model4(3)&KDJ'])
bb=float(mydict2['model4(3)&KDJ'])
cc=float(mydict3['model4(3)&KDJ'])
sheet.cell(row=p+1, column=12).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,13)
sheet.cell(row=p+1, column=13).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,14)
sheet.cell(row=p+1, column=14).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,15)
sheet.cell(row=p+1, column=15).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,16)
KDJ=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,16)
if((KDJ<-10)or(KDJ>110)):
    sign=1
elif((KDJ<55)and(KDJ>45)):
    sign=0.2
else:
    sign=0.95
sheet.cell(row=p+1, column=16).value=sign
if (price<=pricemean):
    amount=(pricemean-price)**3 * 1550 * sign
    biaozhi=1
else:
    amount=(price-pricemean)**3 * -1550 * sign
    biaozhi=0
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=5).value=amount/price
sheet.cell(row=p+1, column=6).value=amount/price+float(aa)
sheet.cell(row=p+1, column=7).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+4).font=fontText1
    sheet.cell(row=p+1, column=i+4).number_format = '0.00_);-0.00'
for i in range(4):
    sheet.cell(row=p+1, column=i+12).font=fontText1
    sheet.cell(row=p+1, column=i+12).number_format = '0.00_);-0.00'   
sheet.cell(row=p+1, column=15).font=fontText1
sheet.cell(row=p+1, column=15).number_format = '0.00_ '  
sheet.cell(row=p+1, column=16).font=fontText2
sheet.cell(row=p+1, column=16).number_format = 'General'

wb.save(filename5)

