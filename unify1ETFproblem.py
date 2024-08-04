#!/usr/bin/python3
# coding: utf-8

# In[15]:


import openpyxl
import xlrd
from openpyxl.styles import Font,Alignment
import publai1
pub=publai1.publai1()
datelai=pub.date
a=[]
b=[]
c=[]
d=[]
fontText1 = Font(name='Tahoma', size=10, color='ff0000')
fontText2 = Font(name='Tahoma', size=10)
AlignmentText1 = Alignment(horizontal='center')
wb0 = xlrd.open_workbook(r'./lai/valuationquan/HwabaoWPszseinnovation100ETF.xlsx')
##选择要操作的工作表， 返回工作表对象
sheet0 = wb0.sheet_by_name('HwabaoWP_szse_innovation_100')
xx=sheet0.nrows 
price=sheet0.cell_value(xx-1,4)
CCI=sheet0.cell_value(xx-1,20)
high=max(sheet0.cell_value(xx-1,2),sheet0.cell_value(xx-2,13))
low=min(sheet0.cell_value(xx-1,3),sheet0.cell_value(xx-2,14))
filedir3=['szseinnovation100indexmodel1prac.xlsx']
dirba0="./lai/valuationquan/szseinnovation100index"
filename3t=dirba0+"/"+filedir3[0]
data1=xlrd.open_workbook(filename3t)#openpyxl模块读不出数据，只能读出单元格公式
sheet2_data=data1.sheet_by_name('myPEPB')
nrow=sheet2_data.nrows
filedir3=['szseinnovation100ETFmodel1.xlsx','szseinnovation100ETFmodel4(SAR).xlsx','szseinnovation100ETFmodel4(CCI).xlsx','szseinnovation100ETFmodel4(RSI&KDJ).xlsx']
dirba1="./lai/valuationquan/szseinnovation100ETF"
filename3=dirba1+"/"+filedir3[0]
data=xlrd.open_workbook(filename3)#openpyxl模块读不出数据，只能读出单元格公式
wb = openpyxl.load_workbook(filename3)


# In[16]:


for sheet in wb:
    print(sheet.title)
    p=sheet.max_row
    sheet.cell(row=p+1, column=1).value=datelai
    sheet.cell(row=p+1, column=1).alignment=AlignmentText1
    sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
    sheet.cell(row=p+1, column=1).font=fontText1
    sheet.cell(row=p+1, column=2).value=price
    sheet.cell(row=p+1, column=2).font=fontText1
    sheet.cell(row=p+1, column=2).number_format = '0.00000'
    if(sheet.title=='model1&pe'):
        a.append(sheet.title)
        b.append(data.sheet_by_name(sheet.title).cell_value(p-2,5))
        c.append(data.sheet_by_name(sheet.title).cell_value(p-2,7))
#         b.append(sheet1_data.cell_value(p-1,5))
#         c.append(sheet1_data.cell_value(p-1,7))
    else:
        a.append(sheet.title)
        b.append(data.sheet_by_name(sheet.title).cell_value(p-2,4))
        c.append(data.sheet_by_name(sheet.title).cell_value(p-2,6))
#         b.append(sheet1_data.cell_value(p-1,4))
#         c.append(sheet1_data.cell_value(p-1,6))        
mydict1=dict(zip(a,b))
mydict2=dict(zip(a,c))
###################################################################
sheet = wb['model1']
sheet.cell(row=p+1, column=3).value=2000
sheet.cell(row=p+1, column=4).value=2000/price
sheet.cell(row=p+1, column=5).value=2000/price+float(mydict1['model1'])
sheet.cell(row=p+1, column=6).value=(2000/price+float(mydict1['model1']))*price
sheet.cell(row=p+1, column=7).value=float(mydict2['model1'])+2000
sheet.cell(row=p+1, column=8).value=(2000/price+float(mydict1['model1']))*price
sheet.cell(row=p+1, column=9).value=(2000/price+float(mydict1['model1']))*price-float(mydict2['model1'])-2000
for i in range(6):
    sheet.cell(row=p+1, column=i+3).font=fontText1
    sheet.cell(row=p+1, column=i+3).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_ '
###################################################################
sheet = wb['model1&pe']
sheet.cell(row=p+1, column=3).value=sheet2_data.cell_value(nrow-1,2)#PE值
sheet.cell(row=p+1, column=3).font=fontText1
sheet.cell(row=p+1, column=3).number_format = 'General'
sheet.cell(row=p+1, column=4).value=2000
sheet.cell(row=p+1, column=4).font=fontText1
sheet.cell(row=p+1, column=4).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=5).value=2000/price
sheet.cell(row=p+1, column=5).font=fontText1
sheet.cell(row=p+1, column=5).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=6).value=2000/price+float(mydict1['model1&pe'])
sheet.cell(row=p+1, column=6).font=fontText1
sheet.cell(row=p+1, column=6).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=7).value=(2000/price+float(mydict1['model1&pe']))*price
sheet.cell(row=p+1, column=7).font=fontText1
sheet.cell(row=p+1, column=7).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=8).value=float(mydict2['model1&pe'])+2000
sheet.cell(row=p+1, column=8).font=fontText1
sheet.cell(row=p+1, column=8).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).value=(2000/price+float(mydict1['model1&pe']))*price
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=10).value=(2000/price+float(mydict1['model1&pe']))*price-float(mydict2['model1&pe'])-2000
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
###########################################################################
sheet = wb['model1&CCI']
sheet.cell(row=p+1, column=10).value=CCI
sheet.cell(row=p+1, column=10).font=fontText2
sheet.cell(row=p+1, column=10).number_format = 'General'
if(float(CCI)<-100):
    sheet.cell(row=p+1, column=3).value=2000
    sheet.cell(row=p+1, column=4).value=2000/price
    sheet.cell(row=p+1, column=5).value=2000/price+float(mydict1['model1&CCI'])
    sheet.cell(row=p+1, column=6).value=(2000/price+float(mydict1['model1&CCI']))*price
    sheet.cell(row=p+1, column=7).value=2000+float(mydict2['model1&CCI'])
    sheet.cell(row=p+1, column=8).value=(2000/price+float(mydict1['model1&CCI']))*price
    sheet.cell(row=p+1, column=9).value=(2000/price+float(mydict1['model1&CCI']))*price-float(mydict2['model1&CCI'])-2000
else:
    sheet.cell(row=p+1, column=3).value=0
    sheet.cell(row=p+1, column=4).value=0
    sheet.cell(row=p+1, column=5).value=float(mydict1['model1&CCI'])
    sheet.cell(row=p+1, column=6).value=float(mydict1['model1&CCI'])*price
    sheet.cell(row=p+1, column=7).value=float(mydict2['model1&CCI'])
    sheet.cell(row=p+1, column=8).value=float(mydict1['model1&CCI'])*price
    sheet.cell(row=p+1, column=9).value=float(mydict1['model1&CCI'])*price-float(mydict2['model1&CCI'])
for i in range(6):
    sheet.cell(row=p+1, column=i+3).font=fontText1
    sheet.cell(row=p+1, column=i+3).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_ '
###########################################################################
sheet = wb['model1&RSI']
tt=price-float(data.sheet_by_name(sheet.title).cell_value(p-2,1))
sheet.cell(row=p+1, column=10).value=max(tt,0)
sheet.cell(row=p+1, column=11).value=(max(tt,0)+float(data.sheet_by_name(sheet.title).cell_value(p-2,10))*5)/6
sheet.cell(row=p+1, column=12).value=abs(tt)
sheet.cell(row=p+1, column=13).value=(abs(tt)+float(data.sheet_by_name(sheet.title).cell_value(p-2,12))*5)/6
sma1=(max(tt,0)+float(data.sheet_by_name(sheet.title).cell_value(p-2,10))*5)/6
sma2=(abs(tt)+float(data.sheet_by_name(sheet.title).cell_value(p-2,12))*5)/6
sheet.cell(row=p+1, column=14).value=sma1*100/sma2
for i in range(5):
    sheet.cell(row=p+1, column=i+10).font=fontText2
    sheet.cell(row=p+1, column=i+10).number_format = 'General'
if((sma1*100/sma2)<20):
    sheet.cell(row=p+1, column=3).value=3000
    sheet.cell(row=p+1, column=4).value=3000/price
    sheet.cell(row=p+1, column=5).value=3000/price+float(mydict1['model1&RSI'])
    sheet.cell(row=p+1, column=6).value=(3000/price+float(mydict1['model1&RSI']))*price
    sheet.cell(row=p+1, column=7).value=3000+float(mydict2['model1&RSI'])
    sheet.cell(row=p+1, column=8).value=(3000/price+float(mydict1['model1&RSI']))*price
    sheet.cell(row=p+1, column=9).value=(3000/price+float(mydict1['model1&RSI']))*price-float(mydict2['model1&RSI'])-3000
else:
    sheet.cell(row=p+1, column=3).value=2000
    sheet.cell(row=p+1, column=4).value=2000/price
    sheet.cell(row=p+1, column=5).value=2000/price+float(mydict1['model1&RSI'])
    sheet.cell(row=p+1, column=6).value=(2000/price+float(mydict1['model1&RSI']))*price
    sheet.cell(row=p+1, column=7).value=2000+float(mydict2['model1&RSI'])
    sheet.cell(row=p+1, column=8).value=(2000/price+float(mydict1['model1&RSI']))*price
    sheet.cell(row=p+1, column=9).value=(2000/price+float(mydict1['model1&RSI']))*price-float(mydict2['model1&RSI'])-2000
for i in range(6):
    sheet.cell(row=p+1, column=i+3).font=fontText1
    sheet.cell(row=p+1, column=i+3).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_ '
#######################################################################
sheet = wb['model1&KDJ']
sheet.cell(row=p+1, column=10).value=high
sheet.cell(row=p+1, column=11).value=low
maxlai=high
minlai=low
t=0
while t<8:
    maxlai=max(maxlai,data.sheet_by_name(sheet.title).cell_value(p-2-t,9))
    minlai=min(minlai,data.sheet_by_name(sheet.title).cell_value(p-2-t,10))
    t=t+1
print("laienwei")
print(maxlai)
print(minlai)
sheet.cell(row=p+1, column=12).value=maxlai
sheet.cell(row=p+1, column=13).value=minlai
sheet.cell(row=p+1, column=14).value=(price-minlai)*100/(maxlai-minlai)
K=(2*data.sheet_by_name(sheet.title).cell_value(p-2,14)+(price-minlai)*100/(maxlai-minlai))/3
D=(2*data.sheet_by_name(sheet.title).cell_value(p-2,15)+K)/3
J=3*K-2*D
sheet.cell(row=p+1, column=15).value=K
sheet.cell(row=p+1, column=16).value=D
sheet.cell(row=p+1, column=17).value=J
for i in range(8):
    sheet.cell(row=p+1, column=10+i).font=fontText2
    sheet.cell(row=p+1, column=10+i).number_format = 'General'
if(J<0):
    sheet.cell(row=p+1, column=3).value=3000
    sheet.cell(row=p+1, column=4).value=3000/price
    sheet.cell(row=p+1, column=5).value=3000/price+float(mydict1['model1&KDJ'])
    sheet.cell(row=p+1, column=6).value=(3000/price+float(mydict1['model1&KDJ']))*price
    sheet.cell(row=p+1, column=7).value=3000+float(mydict2['model1&KDJ'])  
    sheet.cell(row=p+1, column=8).value=(3000/price+float(mydict1['model1&KDJ']))*price
    sheet.cell(row=p+1, column=9).value=(3000/price+float(mydict1['model1&KDJ']))*price-float(mydict2['model1&KDJ'])-3000
else:
    sheet.cell(row=p+1, column=3).value=2000
    sheet.cell(row=p+1, column=4).value=2000/price
    sheet.cell(row=p+1, column=5).value=2000/price+float(mydict1['model1&KDJ'])
    sheet.cell(row=p+1, column=6).value=(2000/price+float(mydict1['model1&KDJ']))*price
    sheet.cell(row=p+1, column=7).value=2000+float(mydict2['model1&KDJ'])
    sheet.cell(row=p+1, column=8).value=(2000/price+float(mydict1['model1&KDJ']))*price
    sheet.cell(row=p+1, column=9).value=(2000/price+float(mydict1['model1&KDJ']))*price-float(mydict2['model1&KDJ'])-2000
for i in range(6):
    sheet.cell(row=p+1, column=i+3).font=fontText1
    sheet.cell(row=p+1, column=i+3).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_ '
wb.save(filename3)


# In[6]:


filename4=dirba1+"/"+filedir3[1]
data=xlrd.open_workbook(filename4)#openpyxl模块读不出数据，只能读出单元格公式
wb = openpyxl.load_workbook(filename4)
sheet1_data=data.sheet_by_name('model4(1)&SAR_manual_oper')
sheet = wb['model4(1)&SAR_manual_oper']
p=sheet.max_row
print(p)
print(sheet1_data.cell_value(p-2,12))
SAR=sheet1_data.cell_value(p-2,12)

print(sheet1_data.cell_value(p-2,6))
aa=sheet1_data.cell_value(p-2,6)
print(sheet1_data.cell_value(p-2,8))
bb=sheet1_data.cell_value(p-2,8)
print(sheet1_data.cell_value(p-2,11))
cc=sheet1_data.cell_value(p-2,11)
if(SAR<1):
    if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
        amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * 3950
        biaozhi=1
    else:
        amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * -3950
        biaozhi=0
else:
    if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
        amount=2*(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * 3950
        biaozhi=1
    else:
        amount=2*(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * -3950
        biaozhi=0
sheet.cell(row=p+1, column=1).value=datelai
sheet.cell(row=p+1, column=1).alignment=AlignmentText1
sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
sheet.cell(row=p+1, column=1).font=fontText1
sheet.cell(row=p+1, column=2).value=price
sheet.cell(row=p+1, column=2).font=fontText1
sheet.cell(row=p+1, column=2).number_format = '0.00000'
sheet.cell(row=p+1, column=3).value=sheet2_data.cell_value(nrow-1,2)#PE值
sheet.cell(row=p+1, column=3).font=fontText1
sheet.cell(row=p+1, column=3).number_format = 'General'
sheet.cell(row=p+1, column=4).value=sheet2_data.cell_value(nrow-1,3)#PE均值
sheet.cell(row=p+1, column=4).font=fontText1
sheet.cell(row=p+1, column=4).number_format = 'General'
sheet.cell(row=p+1, column=5).value=amount
sheet.cell(row=p+1, column=6).value=amount/price
sheet.cell(row=p+1, column=7).value=amount/price+float(aa)
sheet.cell(row=p+1, column=8).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=9).value=float(bb)+amount
    sheet.cell(row=p+1, column=12).value=float(cc)
else:
    sheet.cell(row=p+1, column=9).value=float(bb)
    sheet.cell(row=p+1, column=12).value=float(cc)-amount
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=8).value+sheet.cell(row=p+1, column=12).value
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=10).value-sheet.cell(row=p+1, column=9).value
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_ '
sheet.cell(row=p+1, column=13).value=SAR
sheet.cell(row=p+1, column=13).font=fontText2
sheet.cell(row=p+1, column=13).number_format = 'General'
for i in range(6):
    sheet.cell(row=p+1, column=i+5).font=fontText1
    sheet.cell(row=p+1, column=i+5).number_format = '0.00_);-0.00'
wb.save(filename4)


# In[27]:

a=[]
b=[]
c=[]
d=[]
filename5=dirba1+"/"+filedir3[2]
data=xlrd.open_workbook(filename5)#openpyxl模块读不出数据，只能读出单元格公式
wb = openpyxl.load_workbook(filename5)
sheet = wb['model4(1)']
p=sheet.max_row
print(p)
for sheet in wb:
    print(sheet.title)
    p=sheet.max_row
    sheet.cell(row=p+1, column=1).value=datelai
    sheet.cell(row=p+1, column=1).alignment=AlignmentText1
    sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
    sheet.cell(row=p+1, column=1).font=fontText1
    sheet.cell(row=p+1, column=2).value=price
    sheet.cell(row=p+1, column=2).font=fontText1
    sheet.cell(row=p+1, column=2).number_format = '0.00000'
    sheet.cell(row=p+1, column=3).value=sheet2_data.cell_value(nrow-1,2)#PE值
    sheet.cell(row=p+1, column=3).font=fontText1
    sheet.cell(row=p+1, column=3).number_format = 'General'
    sheet.cell(row=p+1, column=4).value=sheet2_data.cell_value(nrow-1,3)#PE均值
    sheet.cell(row=p+1, column=4).font=fontText1
    sheet.cell(row=p+1, column=4).number_format = 'General'
    if(sheet.title!='model4(3)vol&CCI_per_day'):
        a.append(sheet.title)
        b.append(data.sheet_by_name(sheet.title).cell_value(p-2,6))
        c.append(data.sheet_by_name(sheet.title).cell_value(p-2,8))
        d.append(data.sheet_by_name(sheet.title).cell_value(p-2,11))
#         b.append(sheet1_data.cell_value(p-1,5))
#         c.append(sheet1_data.cell_value(p-1,7))
    else:
        a.append(sheet.title)
        b.append(data.sheet_by_name(sheet.title).cell_value(p-2,8))
        c.append(data.sheet_by_name(sheet.title).cell_value(p-2,10))
        d.append(data.sheet_by_name(sheet.title).cell_value(p-2,13))
#         b.append(sheet1_data.cell_value(p-1,4))
#         c.append(sheet1_data.cell_value(p-1,6))        
mydict1=dict(zip(a,b))
mydict2=dict(zip(a,c))
mydict3=dict(zip(a,d))
print(mydict1)
print(mydict2)
print(mydict3)
#########################################################################
sheet = wb['model4(1)']
aa=float(mydict1['model4(1)'])
bb=float(mydict2['model4(1)'])
cc=float(mydict3['model4(1)'])
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * 3950
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * -3950
    biaozhi=0
sheet.cell(row=p+1, column=5).value=amount
sheet.cell(row=p+1, column=6).value=amount/price
sheet.cell(row=p+1, column=7).value=amount/price+float(aa)
sheet.cell(row=p+1, column=8).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=9).value=float(bb)+amount
    sheet.cell(row=p+1, column=12).value=float(cc)
else:
    sheet.cell(row=p+1, column=9).value=float(bb)
    sheet.cell(row=p+1, column=12).value=float(cc)-amount
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=8).value+sheet.cell(row=p+1, column=12).value
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=10).value-sheet.cell(row=p+1, column=9).value
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+5).font=fontText1
    sheet.cell(row=p+1, column=i+5).number_format = '0.00_);-0.00'
############################################################################
sheet = wb['model4(1)&CCI_per_day']
aa=float(mydict1['model4(1)&CCI_per_day'])
bb=float(mydict2['model4(1)&CCI_per_day'])
cc=float(mydict3['model4(1)&CCI_per_day'])
CCI=sheet0.cell_value(xx-1,20)
sheet.cell(row=p+1, column=13).value=CCI
sheet.cell(row=p+1, column=13).font=fontText2
sheet.cell(row=p+1, column=13).number_format = 'General'
if (sheet2_data.cell_value(nrow-1,2)<sheet2_data.cell_value(nrow-1,3))and(CCI<-100):
    sign=1.2
elif (sheet2_data.cell_value(nrow-1,2)>sheet2_data.cell_value(nrow-1,3))and(CCI>100):
    sign=1.2
else:
    sign=1
sheet.cell(row=p+1, column=14).value=sign
sheet.cell(row=p+1, column=14).font=fontText2
sheet.cell(row=p+1, column=14).number_format = 'General'
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * 3950 * sign
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * -3950 * sign
    biaozhi=0
sheet.cell(row=p+1, column=5).value=amount
sheet.cell(row=p+1, column=6).value=amount/price
sheet.cell(row=p+1, column=7).value=amount/price+float(aa)
sheet.cell(row=p+1, column=8).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=9).value=float(bb)+amount
    sheet.cell(row=p+1, column=12).value=float(cc)
else:
    sheet.cell(row=p+1, column=9).value=float(bb)
    sheet.cell(row=p+1, column=12).value=float(cc)-amount
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=8).value+sheet.cell(row=p+1, column=12).value
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=10).value-sheet.cell(row=p+1, column=9).value
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+5).font=fontText1
    sheet.cell(row=p+1, column=i+5).number_format = '0.00_);-0.00'
############################################################################
sheet = wb['model4(3)&CCI_per_day']
aa=float(mydict1['model4(3)&CCI_per_day'])
bb=float(mydict2['model4(3)&CCI_per_day'])
cc=float(mydict3['model4(3)&CCI_per_day'])
CCI=sheet0.cell_value(xx-1,20)
sheet.cell(row=p+1, column=13).value=CCI
sheet.cell(row=p+1, column=13).font=fontText2
sheet.cell(row=p+1, column=13).number_format = 'General'
if (sheet2_data.cell_value(nrow-1,2)<sheet2_data.cell_value(nrow-1,3))and(CCI<-100):
    sign=1.2
elif (sheet2_data.cell_value(nrow-1,2)>sheet2_data.cell_value(nrow-1,3))and(CCI>100):
    sign=1.2
else:
    sign=1
sheet.cell(row=p+1, column=14).value=sign
sheet.cell(row=p+1, column=14).font=fontText2
sheet.cell(row=p+1, column=14).number_format = 'General'
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 3 * 3950 * sign
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 3 * -3950 * sign
    biaozhi=0
sheet.cell(row=p+1, column=5).value=amount
sheet.cell(row=p+1, column=6).value=amount/price
sheet.cell(row=p+1, column=7).value=amount/price+float(aa)
sheet.cell(row=p+1, column=8).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=9).value=float(bb)+amount
    sheet.cell(row=p+1, column=12).value=float(cc)
else:
    sheet.cell(row=p+1, column=9).value=float(bb)
    sheet.cell(row=p+1, column=12).value=float(cc)-amount
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=8).value+sheet.cell(row=p+1, column=12).value
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=10).value-sheet.cell(row=p+1, column=9).value
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+5).font=fontText1
    sheet.cell(row=p+1, column=i+5).number_format = '0.00_);-0.00'
############################################################################
sheet = wb['model4(3)vol&CCI_per_day']
aa=float(mydict1['model4(3)vol&CCI_per_day'])
bb=float(mydict2['model4(3)vol&CCI_per_day'])
cc=float(mydict3['model4(3)vol&CCI_per_day'])
CCI=sheet0.cell_value(xx-1,20)
sheet.cell(row=p+1, column=15).value=CCI
sheet.cell(row=p+1, column=15).font=fontText2
sheet.cell(row=p+1, column=15).number_format = 'General'
if (sheet2_data.cell_value(nrow-1,2)<sheet2_data.cell_value(nrow-1,3))and(CCI<-100):
    sign=1.2
elif (sheet2_data.cell_value(nrow-1,2)>sheet2_data.cell_value(nrow-1,3))and(CCI>100):
    sign=1.2
else:
    sign=1
sheet.cell(row=p+1, column=16).value=sign
sheet.cell(row=p+1, column=16).font=fontText2
sheet.cell(row=p+1, column=16).number_format = 'General'
vol=sheet0.cell_value(xx-1,5)
volmean=sheet0.cell_value(xx-1,8)
print(type(volmean))
print(volmean)
sheet.cell(row=p+1, column=5).value=vol
sheet.cell(row=p+1, column=5).font=fontText1
sheet.cell(row=p+1, column=5).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=6).value=volmean
sheet.cell(row=p+1, column=6).font=fontText1
sheet.cell(row=p+1, column=6).number_format = '0.00_);-0.00'
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 3 * 3950 * sign * vol / volmean
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 3 * -3950 * sign * vol / volmean
    biaozhi=0
sheet.cell(row=p+1, column=7).value=amount
sheet.cell(row=p+1, column=8).value=amount/price
sheet.cell(row=p+1, column=9).value=amount/price+float(aa)
sheet.cell(row=p+1, column=10).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=11).value=float(bb)+amount
    sheet.cell(row=p+1, column=14).value=float(cc)
else:
    sheet.cell(row=p+1, column=11).value=float(bb)
    sheet.cell(row=p+1, column=14).value=float(cc)-amount
sheet.cell(row=p+1, column=14).font=fontText1
sheet.cell(row=p+1, column=14).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=12).value=sheet.cell(row=p+1, column=10).value+sheet.cell(row=p+1, column=14).value
sheet.cell(row=p+1, column=13).value=sheet.cell(row=p+1, column=12).value-sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=13).font=fontText1
sheet.cell(row=p+1, column=13).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+7).font=fontText1
    sheet.cell(row=p+1, column=i+7).number_format = '0.00_);-0.00'
############################################################################

sheet = wb['model4(1)&CCI_per_month']
aa=float(mydict1['model4(1)&CCI_per_month'])
bb=float(mydict2['model4(1)&CCI_per_month'])
cc=float(mydict3['model4(1)&CCI_per_month'])
sheet.cell(row=p+1, column=13).value=high
sheet.cell(row=p+1, column=14).value=low
typ=(price+low+high)/3
sheet.cell(row=p+1, column=15).value=typ
sumtyp=typ
for i in range(13):
    sumtyp=sumtyp+float(data.sheet_by_name(sheet.title).cell_value(p-2-i,14))
sheet.cell(row=p+1, column=16).value=sumtyp/14
sheet.cell(row=p+1, column=17).value=typ-sumtyp/14
avedev=abs(typ-sumtyp/14)
for i in range(13):
    avedev=avedev+abs(float(data.sheet_by_name(sheet.title).cell_value(p-2-i,14))-sumtyp/14)
sheet.cell(row=p+1, column=18).value=avedev/14
sheet.cell(row=p+1, column=19).value=0.015*avedev/14
CCI=(typ-sumtyp/14)/(0.015*avedev/14)
sheet.cell(row=p+1, column=20).value=CCI
if (sheet2_data.cell_value(nrow-1,2)<sheet2_data.cell_value(nrow-1,3))and(CCI<-100):
    sign=1.2
elif (sheet2_data.cell_value(nrow-1,2)>sheet2_data.cell_value(nrow-1,3))and(CCI>100):
    sign=1.2
else:
    sign=1
sheet.cell(row=p+1, column=21).value=sign
for i in range(4):
    sheet.cell(row=p+1, column=i+13).font=fontText1
    sheet.cell(row=p+1, column=i+13).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=17).font=fontText1
sheet.cell(row=p+1, column=17).number_format = '0.00_ '
sheet.cell(row=p+1, column=18).font=fontText1
sheet.cell(row=p+1, column=18).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=19).font=fontText2
sheet.cell(row=p+1, column=19).number_format = '0.00000000_);(0.00000000)'
sheet.cell(row=p+1, column=20).font=fontText2
sheet.cell(row=p+1, column=20).number_format = '0.000000_ '
sheet.cell(row=p+1, column=21).font=fontText2
sheet.cell(row=p+1, column=21).number_format = 'General'
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * 3950 * sign
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * -3950 * sign
    biaozhi=0
sheet.cell(row=p+1, column=5).value=amount
sheet.cell(row=p+1, column=6).value=amount/price
sheet.cell(row=p+1, column=7).value=amount/price+float(aa)
sheet.cell(row=p+1, column=8).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=9).value=float(bb)+amount
    sheet.cell(row=p+1, column=12).value=float(cc)
else:
    sheet.cell(row=p+1, column=9).value=float(bb)
    sheet.cell(row=p+1, column=12).value=float(cc)-amount
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=8).value+sheet.cell(row=p+1, column=12).value
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=10).value-sheet.cell(row=p+1, column=9).value
for i in range(6):
    sheet.cell(row=p+1, column=i+5).font=fontText1
    sheet.cell(row=p+1, column=i+5).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_ '
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_);-0.00'
wb.save(filename5)


# In[17]:


a=[]
b=[]
c=[]
d=[]
filename6=dirba1+"/"+filedir3[3]
data=xlrd.open_workbook(filename6)#openpyxl模块读不出数据，只能读出单元格公式
wb = openpyxl.load_workbook(filename6)
sheet = wb['model4(1)']
p=sheet.max_row
print(p)
for sheet in wb:
    print(sheet.title)
    p=sheet.max_row
    sheet.cell(row=p+1, column=1).value=datelai
    sheet.cell(row=p+1, column=1).alignment=AlignmentText1
    sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
    sheet.cell(row=p+1, column=1).font=fontText1
    sheet.cell(row=p+1, column=2).value=price
    sheet.cell(row=p+1, column=2).font=fontText1
    sheet.cell(row=p+1, column=2).number_format = '0.00000'
    sheet.cell(row=p+1, column=3).value=sheet2_data.cell_value(nrow-1,2)#PE值
    sheet.cell(row=p+1, column=3).font=fontText1
    sheet.cell(row=p+1, column=3).number_format = 'General'
    sheet.cell(row=p+1, column=4).value=sheet2_data.cell_value(nrow-1,3)#PE均值
    sheet.cell(row=p+1, column=4).font=fontText1
    sheet.cell(row=p+1, column=4).number_format = 'General'
    if(not((sheet.title=='model4(3)vol&RSI')or(sheet.title=='model4(3)vol'))):
        a.append(sheet.title)
        b.append(data.sheet_by_name(sheet.title).cell_value(p-2,6))
        c.append(data.sheet_by_name(sheet.title).cell_value(p-2,8))
        d.append(data.sheet_by_name(sheet.title).cell_value(p-2,11))
#         b.append(sheet1_data.cell_value(p-1,5))
#         c.append(sheet1_data.cell_value(p-1,7))
    else:
        a.append(sheet.title)
        b.append(data.sheet_by_name(sheet.title).cell_value(p-2,8))
        c.append(data.sheet_by_name(sheet.title).cell_value(p-2,10))
        d.append(data.sheet_by_name(sheet.title).cell_value(p-2,13))
#         b.append(sheet1_data.cell_value(p-1,4))
#         c.append(sheet1_data.cell_value(p-1,6))        
mydict1=dict(zip(a,b))
mydict2=dict(zip(a,c))
mydict3=dict(zip(a,d))
print(mydict1)
print(mydict2)
print(mydict3)
#########################################################################
sheet = wb['model4(1)']
aa=float(mydict1['model4(1)'])
bb=float(mydict2['model4(1)'])
cc=float(mydict3['model4(1)'])
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * 3950
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * -3950
    biaozhi=0
sheet.cell(row=p+1, column=5).value=amount
sheet.cell(row=p+1, column=6).value=amount/price
sheet.cell(row=p+1, column=7).value=amount/price+float(aa)
sheet.cell(row=p+1, column=8).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=9).value=float(bb)+amount
    sheet.cell(row=p+1, column=12).value=float(cc)
else:
    sheet.cell(row=p+1, column=9).value=float(bb)
    sheet.cell(row=p+1, column=12).value=float(cc)-amount
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=8).value+sheet.cell(row=p+1, column=12).value
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=10).value-sheet.cell(row=p+1, column=9).value
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+5).font=fontText1
    sheet.cell(row=p+1, column=i+5).number_format = '0.00_);-0.00'
#########################################################################
sheet = wb['model4(3)']
aa=float(mydict1['model4(3)'])
bb=float(mydict2['model4(3)'])
cc=float(mydict3['model4(3)'])
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 3 * 3950
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 3 * -3950
    biaozhi=0
sheet.cell(row=p+1, column=5).value=amount
sheet.cell(row=p+1, column=6).value=amount/price
sheet.cell(row=p+1, column=7).value=amount/price+float(aa)
sheet.cell(row=p+1, column=8).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=9).value=float(bb)+amount
    sheet.cell(row=p+1, column=12).value=float(cc)
else:
    sheet.cell(row=p+1, column=9).value=float(bb)
    sheet.cell(row=p+1, column=12).value=float(cc)-amount
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=8).value+sheet.cell(row=p+1, column=12).value
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=10).value-sheet.cell(row=p+1, column=9).value
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+5).font=fontText1
    sheet.cell(row=p+1, column=i+5).number_format = '0.00_);-0.00'
#########################################################################
sheet = wb['model4(1)&RSI']
aa=float(mydict1['model4(1)&RSI'])
bb=float(mydict2['model4(1)&RSI'])
cc=float(mydict3['model4(1)&RSI'])
datax=xlrd.open_workbook(filename3)#openpyxl模块读不出数据，只能读出单元格公式
sheetx=datax.sheet_by_name('model1&RSI')
nrowx=sheetx.nrows
sheet.cell(row=p+1, column=13).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,10)
sheet.cell(row=p+1, column=14).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,12)
sheet.cell(row=p+1, column=15).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
RSIpre=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
if(float(RSIpre)<20):
    sign=2
elif((float(RSIpre)>20)and(RSIpre<25)):
    sign=1
elif((float(RSIpre)>25)and(RSIpre<50)):
    sign=0.95
elif((float(RSIpre)>50)and(RSIpre<80)):
    sign=0.2
else:
    sign=1
sheet.cell(row=p+1, column=16).value=sign
sheet.cell(row=p+1, column=16).font=fontText2
sheet.cell(row=p+1, column=16).number_format = 'General'
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * 3950 * sign
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * -3950 * sign
    biaozhi=0
sheet.cell(row=p+1, column=5).value=amount
sheet.cell(row=p+1, column=6).value=amount/price
sheet.cell(row=p+1, column=7).value=amount/price+float(aa)
sheet.cell(row=p+1, column=8).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=9).value=float(bb)+amount
    sheet.cell(row=p+1, column=12).value=float(cc)
else:
    sheet.cell(row=p+1, column=9).value=float(bb)
    sheet.cell(row=p+1, column=12).value=float(cc)-amount
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=8).value+sheet.cell(row=p+1, column=12).value
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=10).value-sheet.cell(row=p+1, column=9).value
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+5).font=fontText1
    sheet.cell(row=p+1, column=i+5).number_format = '0.00_);-0.00'
for i in range(3):
    sheet.cell(row=p+1, column=i+13).font=fontText1
    sheet.cell(row=p+1, column=i+13).number_format = '0.000000_);(0.000000)' 
#########################################################################
sheet = wb['model4(3)&RSI']
aa=float(mydict1['model4(3)&RSI'])
bb=float(mydict2['model4(3)&RSI'])
cc=float(mydict3['model4(3)&RSI'])
sheet.cell(row=p+1, column=13).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,10)
sheet.cell(row=p+1, column=14).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,12)
sheet.cell(row=p+1, column=15).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
RSIpre=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
if(float(RSIpre)<20):
    sign=2
elif((float(RSIpre)>20)and(RSIpre<25)):
    sign=1
elif((float(RSIpre)>25)and(RSIpre<50)):
    sign=0.95
elif((float(RSIpre)>50)and(RSIpre<80)):
    sign=0.2
else:
    sign=1
sheet.cell(row=p+1, column=16).value=sign
sheet.cell(row=p+1, column=16).font=fontText2
sheet.cell(row=p+1, column=16).number_format = 'General'
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 3 * 3950 * sign
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 3 * -3950 * sign
    biaozhi=0
sheet.cell(row=p+1, column=5).value=amount
sheet.cell(row=p+1, column=6).value=amount/price
sheet.cell(row=p+1, column=7).value=amount/price+float(aa)
sheet.cell(row=p+1, column=8).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=9).value=float(bb)+amount
    sheet.cell(row=p+1, column=12).value=float(cc)
else:
    sheet.cell(row=p+1, column=9).value=float(bb)
    sheet.cell(row=p+1, column=12).value=float(cc)-amount
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=8).value+sheet.cell(row=p+1, column=12).value
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=10).value-sheet.cell(row=p+1, column=9).value
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+5).font=fontText1
    sheet.cell(row=p+1, column=i+5).number_format = '0.00_);-0.00'
for i in range(3):
    sheet.cell(row=p+1, column=i+13).font=fontText1
    sheet.cell(row=p+1, column=i+13).number_format = '0.000000_);(0.000000)' 
#########################################################################
sheet = wb['model4(1)&KDJ']
aa=float(mydict1['model4(1)&KDJ'])
bb=float(mydict2['model4(1)&KDJ'])
cc=float(mydict3['model4(1)&KDJ'])
sheet.cell(row=p+1, column=13).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,13)
sheet.cell(row=p+1, column=14).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,14)
sheet.cell(row=p+1, column=15).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,15)
sheet.cell(row=p+1, column=16).value=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,16)
KDJ=datax.sheet_by_name('model1&KDJ').cell_value(nrowx-1,16)
if((KDJ<0)or(KDJ>100)):
    sign=1.2
else:
    sign=1
sheet.cell(row=p+1, column=17).value=sign
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * 3950 * sign
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 2 * -3950 * sign
    biaozhi=0
sheet.cell(row=p+1, column=5).value=amount
sheet.cell(row=p+1, column=6).value=amount/price
sheet.cell(row=p+1, column=7).value=amount/price+float(aa)
sheet.cell(row=p+1, column=8).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=9).value=float(bb)+amount
    sheet.cell(row=p+1, column=12).value=float(cc)
else:
    sheet.cell(row=p+1, column=9).value=float(bb)
    sheet.cell(row=p+1, column=12).value=float(cc)-amount
sheet.cell(row=p+1, column=12).font=fontText1
sheet.cell(row=p+1, column=12).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=8).value+sheet.cell(row=p+1, column=12).value
sheet.cell(row=p+1, column=11).value=sheet.cell(row=p+1, column=10).value-sheet.cell(row=p+1, column=9).value
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_ '
for i in range(6):
    sheet.cell(row=p+1, column=i+5).font=fontText1
    sheet.cell(row=p+1, column=i+5).number_format = '0.00_);-0.00'
for i in range(3):
    sheet.cell(row=p+1, column=i+13).font=fontText1
    sheet.cell(row=p+1, column=i+13).number_format = '0.00_);-0.00'   
sheet.cell(row=p+1, column=16).font=fontText1
sheet.cell(row=p+1, column=16).number_format = '0.00_ '  
sheet.cell(row=p+1, column=17).font=fontText2
sheet.cell(row=p+1, column=17).number_format = 'General'
#########################################################################
sheet = wb['model4(3)vol']
aa=float(mydict1['model4(3)vol'])
bb=float(mydict2['model4(3)vol'])
cc=float(mydict3['model4(3)vol'])
vol=sheet0.cell_value(xx-1,5)
volmean=sheet0.cell_value(xx-1,8)
print(type(volmean))
print(volmean)
sheet.cell(row=p+1, column=5).value=vol
sheet.cell(row=p+1, column=5).font=fontText1
sheet.cell(row=p+1, column=5).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=6).value=volmean
sheet.cell(row=p+1, column=6).font=fontText1
sheet.cell(row=p+1, column=6).number_format = '0.00_);-0.00'
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 3 * 3950 * vol / volmean
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 3 * -3950 * vol / volmean
    biaozhi=0
sheet.cell(row=p+1, column=7).value=amount
sheet.cell(row=p+1, column=8).value=amount/price
sheet.cell(row=p+1, column=9).value=amount/price+float(aa)
sheet.cell(row=p+1, column=10).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=11).value=float(bb)+amount
    sheet.cell(row=p+1, column=14).value=float(cc)
else:
    sheet.cell(row=p+1, column=11).value=float(bb)
    sheet.cell(row=p+1, column=14).value=float(cc)-amount
sheet.cell(row=p+1, column=14).font=fontText1
sheet.cell(row=p+1, column=14).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=12).value=sheet.cell(row=p+1, column=10).value+sheet.cell(row=p+1, column=14).value
sheet.cell(row=p+1, column=13).value=sheet.cell(row=p+1, column=12).value-sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=13).font=fontText1
sheet.cell(row=p+1, column=13).number_format = '0.00_ '
for i in range(8):
    sheet.cell(row=p+1, column=i+5).font=fontText1
    sheet.cell(row=p+1, column=i+5).number_format = '0.00_);-0.00'
#########################################################################
sheet = wb['model4(3)vol&RSI']
aa=float(mydict1['model4(3)vol&RSI'])
bb=float(mydict2['model4(3)vol&RSI'])
cc=float(mydict3['model4(3)vol&RSI'])
sheet.cell(row=p+1, column=15).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,10)
sheet.cell(row=p+1, column=16).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,12)
sheet.cell(row=p+1, column=17).value=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
RSIpre=datax.sheet_by_name('model1&RSI').cell_value(nrowx-1,13)
if(float(RSIpre)<20):
    sign=2
elif((float(RSIpre)>20)and(RSIpre<25)):
    sign=1
elif((float(RSIpre)>25)and(RSIpre<50)):
    sign=0.95
elif((float(RSIpre)>50)and(RSIpre<80)):
    sign=0.2
else:
    sign=1
sheet.cell(row=p+1, column=18).value=sign
sheet.cell(row=p+1, column=18).font=fontText2
sheet.cell(row=p+1, column=18).number_format = 'General'
vol=sheet0.cell_value(xx-1,5)
volmean=sheet0.cell_value(xx-1,8)
print(type(volmean))
print(volmean)
sheet.cell(row=p+1, column=5).value=vol
sheet.cell(row=p+1, column=5).font=fontText1
sheet.cell(row=p+1, column=5).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=6).value=volmean
sheet.cell(row=p+1, column=6).font=fontText1
sheet.cell(row=p+1, column=6).number_format = '0.00_);-0.00'
if (sheet2_data.cell_value(nrow-1,2)<=sheet2_data.cell_value(nrow-1,3)):
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 3 * sign * 550 * vol / volmean
    biaozhi=1
else:
    amount=(sheet2_data.cell_value(nrow-1,3)-sheet2_data.cell_value(nrow-1,2))** 3 * sign * -550 * vol / volmean
    biaozhi=0
sheet.cell(row=p+1, column=7).value=amount
sheet.cell(row=p+1, column=8).value=amount/price
sheet.cell(row=p+1, column=9).value=amount/price+float(aa)
sheet.cell(row=p+1, column=10).value=(amount/price+float(aa))*price
if(biaozhi==1):
    sheet.cell(row=p+1, column=11).value=float(bb)+amount
    sheet.cell(row=p+1, column=14).value=float(cc)
else:
    sheet.cell(row=p+1, column=11).value=float(bb)
    sheet.cell(row=p+1, column=14).value=float(cc)-amount
sheet.cell(row=p+1, column=14).font=fontText1
sheet.cell(row=p+1, column=14).number_format = '0.00_);-0.00'
sheet.cell(row=p+1, column=12).value=sheet.cell(row=p+1, column=10).value+sheet.cell(row=p+1, column=14).value
sheet.cell(row=p+1, column=13).value=sheet.cell(row=p+1, column=12).value-sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=13).font=fontText1
sheet.cell(row=p+1, column=13).number_format = '0.00_ '
for i in range(8):
    sheet.cell(row=p+1, column=i+5).font=fontText1
    sheet.cell(row=p+1, column=i+5).number_format = '0.00_);-0.00'
for i in range(3):
    sheet.cell(row=p+1, column=i+15).font=fontText1
    sheet.cell(row=p+1, column=i+15).number_format = '0.000000_);(0.000000)' 
wb.save(filename6)


# In[10]:







