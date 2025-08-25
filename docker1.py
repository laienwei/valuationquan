#!/usr/bin/python3
# coding: utf-8

import os
import openpyxl
import publai1
pub=publai1.publai1()
with open(pub.filename,"r") as fp:
    lines=fp.readlines()
num=len(lines)
filename=['./lai/valuationquan/szseinnovation100index/szseinnovation100indexmodel1.xlsx','./lai/valuationquan/szseinnovation100index/szseinnovation100indexmodel1cn.xlsx','./lai/valuationquan/CNIESG300index/CNIESG300indexmodel1.xlsx','./lai/valuationquan/CNIESG300index/CNIESG300indexmodel1cn.xlsx']
os.system("cp " + filename[0] +" ../")
os.system("cp " + filename[1] +" ../")
os.system("cp " + filename[2] +" ../")
os.system("cp " + filename[3] +" ../")

def del_ws_rows(wb1,rowd,filena):
    ws1=wb1['myPEPB']
    rowd=sorted(rowd,reverse=True)
    for r in rowd:
        ws1.delete_rows(r)
    wb1.save(filena)
wb = openpyxl.load_workbook(filename[0])
ws=wb['myPEPB']
p=ws.max_row
c=list(range(p-num+1,p+1))
del_ws_rows(wb,c,filename[0])
wb.close()
wb = openpyxl.load_workbook(filename[1])
c=list(range(p-num+1,p+1))
del_ws_rows(wb,c,filename[1])
wb.close()
wb = openpyxl.load_workbook(filename[2])
c=list(range(p-3-num+1,p-3+1))
del_ws_rows(wb,c,filename[2])
wb.close()
wb = openpyxl.load_workbook(filename[3])
c=list(range(p-3-num+1,p-3+1))
del_ws_rows(wb,c,filename[3])
wb.close()
#####################################################
os.system("cp " + ".cnb.yml " + ".cnb3.yml")
os.system("mv " + ".cnb2.yml " + ".cnb.yml")
os.system("mv " + ".cnb3.yml " + ".cnb2.yml")
input()
os.system("mv ../" + filename[0][-33:] +" " + filename[0][0:43])
os.system("mv ../" + filename[1][-35:] +" " + filename[1][0:43])

os.system("mv ../" + filename[2][-25:] +" " + filename[2][0:35])
os.system("mv ../" + filename[3][-27:] +" " + filename[3][0:35])
#####################################################
os.system("cp " + ".cnb.yml " + ".cnb3.yml")
os.system("mv " + ".cnb2.yml " + ".cnb.yml")
os.system("mv " + ".cnb3.yml " + ".cnb2.yml")