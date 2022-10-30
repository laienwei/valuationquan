#!/usr/bin/python3
# coding: utf-8

import os
import openpyxl as ox
file_dir=r'./date'
laistr='result'
laistr1=".csv"
laistr11="Jing.csv"
laistr12="Dong.csv"
laistr13="PB.csv"
year="2022"
for root,dirs,files in os.walk(file_dir):
    for q in range(len(dirs)):
        print(dirs[q])
        star=dirs[q]
        file_dir1=file_dir+'/'+star+'/'+laistr+star+laistr11
        with open(file_dir1,"r") as fp:
             lines=fp.readlines()
        num2=len(lines)
        for i in range(num2):
            lines[i]=lines[i][:-1]
        PEjing=lines
        lines=[]
        file_dir1=file_dir+'/'+star+'/'+laistr+star+laistr12
        with open(file_dir1,"r") as fp:
             lines=fp.readlines()
        for i in range(num2):
            lines[i]=lines[i][:-1]
        PEdong=lines
        lines=[]
        file_dir1=file_dir+'/'+star+'/'+laistr+star+laistr13
        with open(file_dir1,"r") as fp:
             lines=fp.readlines()
        for i in range(num2):
            lines[i]=lines[i][:-1]
        PB=lines
        lines=[]
        if(star[0]=="0"):
            date1=year+"/"+star[1:2]+"/"+star[2:]
            print(date1)
        else:
            date1=year+"/"+star[0:2]+"/"+star[2:]
            print(date1)
#         date1="2022/7/25"
        wb=ox.load_workbook(r'./crawler/headpagepestatic.xlsx')
#         mymac=wb.macro('ptxoo1')
        sh=wb['start_from_2021.4.9']
        lairow=sh.max_row+1
        sh.cell(row=lairow,column=1,value=date1)        
        for npp in range(num2):
            sh.cell(row=lairow,column=npp+2,value=float(PEjing[num2-1-npp]))
        wb.save(r'./crawler/headpagepestatic.xlsx')
        wb=ox.load_workbook(r'./crawler/headpagepedynamic.xlsx')
        sh=wb['start_from_2021.4.9']
        lairow=sh.max_row+1
        sh.cell(row=lairow,column=1,value=date1)        
        for npp in range(num2):
            sh.cell(row=lairow,column=npp+2,value=float(PEdong[num2-1-npp]))
        wb.save(r'./crawler/headpagepedynamic.xlsx')
        wb=ox.load_workbook(r'./crawler/headpagepb.xlsx')
        sh=wb['start_from_2021.4.9']
        lairow=sh.max_row+1
        sh.cell(row=lairow,column=1,value=date1)        
        for npp in range(num2):
            sh.cell(row=lairow,column=npp+2,value=float(PB[num2-1-npp]))
        wb.save(r'./crawler/headpagepb.xlsx')
########################################################################################
        file_dir1=file_dir+'/'+star+'/'+laistr+star+laistr1
        with open(file_dir1,"r") as fp:
             lines=fp.readlines()
        num1=len(lines)
        for i in range(num1):
            lines[i]=lines[i][:-1]
        n=lines
        print(num1)
        t=0
        wb=ox.load_workbook(r'./crawler/newworkbook.xlsx')
        sh=wb['start_from_2021.4.9&2021.4.6']
        lairow=sh.max_row+1
        sh.cell(row=lairow,column=1,value=date1)   
#         wb.save(r' ./crawler/newworkbook.xlsm')
        while t<56:
             sh.cell(row=lairow,column=t*5+2,value=float(n[t*5]))
             sh.cell(row=lairow,column=t*5+3,value=float(n[t*5+1]))
             sh.cell(row=lairow,column=t*5+4,value=float(n[t*5+2]))
             sh.cell(row=lairow,column=t*5+5,value=float(n[t*5+3]))
             sh.cell(row=lairow,column=t*5+6,value=float(n[t*5+4]))
             t=t+1
        wb.save(r'./crawler/newworkbook.xlsx')




