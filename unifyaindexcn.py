#!/usr/bin/env python
# coding: utf-8

# In[ ]:


#!/usr/bin/env python
# coding: utf-8

# In[ ]:


#!/usr/bin/python3
# coding: utf-8

# In[28]:
import openpyxl
import xlrd
from openpyxl.styles import Font,Alignment
datelai='2023-04-28'
fontText1 = Font(name='Tahoma', size=10, color='ff0000')
AlignmentText1 = Alignment(horizontal='center')
wb0 = openpyxl.load_workbook(r'./lai/valuationquan/AIinduindex.xlsx')
##选择要操作的工作表， 返回工作表对象
sheet0 = wb0['AI_indu_index']
xx=sheet0.max_row 
index=sheet0.cell(row=xx,column=5).value/1000
indexmean=sheet0.cell(row=xx,column=10).value
indexMA250=sheet0.cell(row=xx,column=11).value
filedir3=['AIinduindexmodel2cn.xlsx','AIinduindexmodel4cn.xlsx']
dirba1="./lai/valuationquan/AIindex"
filename3=dirba1+"/"+filedir3[0]
data=xlrd.open_workbook(filename3)#openpyxl模块读不出数据，只能读出单元格公式

# In[38]:
filename4=dirba1+"/"+filedir3[0]
data=xlrd.open_workbook(filename4)#openpyxl模块读不出数据，只能读出单元格公式
sheet1_data=data.sheet_by_name('模型二 (1)平均线')
#filename3=dirba1+"/"+"szseinnovation100indexmodel1prac.xlsx"
#data1=xlrd.open_workbook(filename3)#openpyxl模块读不出数据，只能读出单元格公式
#sheet2_data=data1.sheet_by_name('myPEPB')
#nrow=sheet2_data.nrows

wb = openpyxl.load_workbook(filename4)
sheet = wb['模型二 (1)平均线']
p=sheet.max_row
print(p)
print(sheet1_data.cell_value(p-1,5))
aa=sheet1_data.cell_value(p-1,5)
print(sheet1_data.cell_value(p-1,7))
bb=sheet1_data.cell_value(p-1,7)
print(sheet1_data.cell_value(p-1,10))
cc=sheet1_data.cell_value(p-1,10)
if (index<=indexmean):
    amount=(indexmean-index) * 1550
    biaozhi=1
else:
    amount=(index-indexmean) * -1550
    biaozhi=0
sheet.cell(row=p+1, column=1).value=datelai
sheet.cell(row=p+1, column=1).alignment=AlignmentText1
sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
sheet.cell(row=p+1, column=1).font=fontText1
sheet.cell(row=p+1, column=2).value=index
sheet.cell(row=p+1, column=2).font=fontText1
sheet.cell(row=p+1, column=2).number_format = '0.00000'
sheet.cell(row=p+1, column=3).value=indexmean
sheet.cell(row=p+1, column=3).font=fontText1
sheet.cell(row=p+1, column=3).number_format = '0.00000'
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=4).font=fontText1
sheet.cell(row=p+1, column=4).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=5).value=amount/index
sheet.cell(row=p+1, column=5).font=fontText1
sheet.cell(row=p+1, column=5).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=6).value=amount/index+float(aa)
sheet.cell(row=p+1, column=6).font=fontText1
sheet.cell(row=p+1, column=6).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=7).value=(amount/index+float(aa))*index
sheet.cell(row=p+1, column=7).font=fontText1
sheet.cell(row=p+1, column=7).number_format = '0.00_);(0.00)'
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=8).font=fontText1
sheet.cell(row=p+1, column=8).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '

sheet1_data=data.sheet_by_name('模型二 (1)MA250')
sheet = wb['模型二 (1)MA250']
p=sheet.max_row
print(p)
print(sheet1_data.cell_value(p-1,5))
aa=sheet1_data.cell_value(p-1,5)
print(sheet1_data.cell_value(p-1,7))
bb=sheet1_data.cell_value(p-1,7)
print(sheet1_data.cell_value(p-1,10))
cc=sheet1_data.cell_value(p-1,10)
if (index<=indexmean):
    amount=(indexMA250-index) * 1550
    biaozhi=1
else:
    amount=(index-indexMA250) * -1550
    biaozhi=0
sheet.cell(row=p+1, column=1).value=datelai
sheet.cell(row=p+1, column=1).alignment=AlignmentText1
sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
sheet.cell(row=p+1, column=1).font=fontText1
sheet.cell(row=p+1, column=2).value=index
sheet.cell(row=p+1, column=2).font=fontText1
sheet.cell(row=p+1, column=2).number_format = '0.00000'
sheet.cell(row=p+1, column=3).value=indexmean
sheet.cell(row=p+1, column=3).font=fontText1
sheet.cell(row=p+1, column=3).number_format = '0.00000'
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=4).font=fontText1
sheet.cell(row=p+1, column=4).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=5).value=amount/index
sheet.cell(row=p+1, column=5).font=fontText1
sheet.cell(row=p+1, column=5).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=6).value=amount/index+float(aa)
sheet.cell(row=p+1, column=6).font=fontText1
sheet.cell(row=p+1, column=6).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=7).value=(amount/index+float(aa))*index
sheet.cell(row=p+1, column=7).font=fontText1
sheet.cell(row=p+1, column=7).number_format = '0.00_);(0.00)'
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=8).font=fontText1
sheet.cell(row=p+1, column=8).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '

wb.save(filename4)

filename5=dirba1+"/"+filedir3[1]
data=xlrd.open_workbook(filename5)#openpyxl模块读不出数据，只能读出单元格公式
############################################
sheet1_data=data.sheet_by_name('模型四 (1)平均线')

wb = openpyxl.load_workbook(filename5)
sheet = wb['模型四 (1)平均线']
p=sheet.max_row
print(p)
print(sheet1_data.cell_value(p-1,5))
aa=sheet1_data.cell_value(p-1,5)
print(sheet1_data.cell_value(p-1,7))
bb=sheet1_data.cell_value(p-1,7)
print(sheet1_data.cell_value(p-1,10))
cc=sheet1_data.cell_value(p-1,10)
if (index<=indexmean):
    amount=(indexmean-index)**2 * 1550
    biaozhi=1
else:
    amount=(index-indexmean)**2 * -1550
    biaozhi=0
sheet.cell(row=p+1, column=1).value=datelai
sheet.cell(row=p+1, column=1).alignment=AlignmentText1
sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
sheet.cell(row=p+1, column=1).font=fontText1
sheet.cell(row=p+1, column=2).value=index
sheet.cell(row=p+1, column=2).font=fontText1
sheet.cell(row=p+1, column=2).number_format = '0.00000'
sheet.cell(row=p+1, column=3).value=indexmean
sheet.cell(row=p+1, column=3).font=fontText1
sheet.cell(row=p+1, column=3).number_format = '0.00000'
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=4).font=fontText1
sheet.cell(row=p+1, column=4).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=5).value=amount/index
sheet.cell(row=p+1, column=5).font=fontText1
sheet.cell(row=p+1, column=5).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=6).value=amount/index+float(aa)
sheet.cell(row=p+1, column=6).font=fontText1
sheet.cell(row=p+1, column=6).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=7).value=(amount/index+float(aa))*index
sheet.cell(row=p+1, column=7).font=fontText1
sheet.cell(row=p+1, column=7).number_format = '0.00_);(0.00)'
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=8).font=fontText1
sheet.cell(row=p+1, column=8).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '

sheet1_data=data.sheet_by_name('模型四 (1)MA250')
sheet = wb['模型四 (1)MA250']
p=sheet.max_row
print(p)
print(sheet1_data.cell_value(p-1,5))
aa=sheet1_data.cell_value(p-1,5)
print(sheet1_data.cell_value(p-1,7))
bb=sheet1_data.cell_value(p-1,7)
print(sheet1_data.cell_value(p-1,10))
cc=sheet1_data.cell_value(p-1,10)
if (index<=indexmean):
    amount=(indexMA250-index)**2 * 1550
    biaozhi=1
else:
    amount=(index-indexMA250)**2 * -1550
    biaozhi=0
sheet.cell(row=p+1, column=1).value=datelai
sheet.cell(row=p+1, column=1).alignment=AlignmentText1
sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
sheet.cell(row=p+1, column=1).font=fontText1
sheet.cell(row=p+1, column=2).value=index
sheet.cell(row=p+1, column=2).font=fontText1
sheet.cell(row=p+1, column=2).number_format = '0.00000'
sheet.cell(row=p+1, column=3).value=indexmean
sheet.cell(row=p+1, column=3).font=fontText1
sheet.cell(row=p+1, column=3).number_format = '0.00000'
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=4).font=fontText1
sheet.cell(row=p+1, column=4).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=5).value=amount/index
sheet.cell(row=p+1, column=5).font=fontText1
sheet.cell(row=p+1, column=5).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=6).value=amount/index+float(aa)
sheet.cell(row=p+1, column=6).font=fontText1
sheet.cell(row=p+1, column=6).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=7).value=(amount/index+float(aa))*index
sheet.cell(row=p+1, column=7).font=fontText1
sheet.cell(row=p+1, column=7).number_format = '0.00_);(0.00)'
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=8).font=fontText1
sheet.cell(row=p+1, column=8).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '

sheet1_data=data.sheet_by_name('模型四 (3)MA250')
sheet = wb['模型四 (3)MA250']
p=sheet.max_row
print(p)
print(sheet1_data.cell_value(p-1,5))
aa=sheet1_data.cell_value(p-1,5)
print(sheet1_data.cell_value(p-1,7))
bb=sheet1_data.cell_value(p-1,7)
print(sheet1_data.cell_value(p-1,10))
cc=sheet1_data.cell_value(p-1,10)
if (index<=indexmean):
    amount=(indexMA250-index)**3 * 1550
    biaozhi=1
else:
    amount=(index-indexMA250)**3 * -1550
    biaozhi=0
sheet.cell(row=p+1, column=1).value=datelai
sheet.cell(row=p+1, column=1).alignment=AlignmentText1
sheet.cell(row=p+1, column=1).number_format = 'yyyy-mm-dd'
sheet.cell(row=p+1, column=1).font=fontText1
sheet.cell(row=p+1, column=2).value=index
sheet.cell(row=p+1, column=2).font=fontText1
sheet.cell(row=p+1, column=2).number_format = '0.00000'
sheet.cell(row=p+1, column=3).value=indexmean
sheet.cell(row=p+1, column=3).font=fontText1
sheet.cell(row=p+1, column=3).number_format = '0.00000'
sheet.cell(row=p+1, column=4).value=amount
sheet.cell(row=p+1, column=4).font=fontText1
sheet.cell(row=p+1, column=4).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=5).value=amount/index
sheet.cell(row=p+1, column=5).font=fontText1
sheet.cell(row=p+1, column=5).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=6).value=amount/index+float(aa)
sheet.cell(row=p+1, column=6).font=fontText1
sheet.cell(row=p+1, column=6).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=7).value=(amount/index+float(aa))*index
sheet.cell(row=p+1, column=7).font=fontText1
sheet.cell(row=p+1, column=7).number_format = '0.00_);(0.00)'
if(biaozhi==1):
    sheet.cell(row=p+1, column=8).value=float(bb)+amount
    sheet.cell(row=p+1, column=11).value=float(cc)
else:
    sheet.cell(row=p+1, column=8).value=float(bb)
    sheet.cell(row=p+1, column=11).value=float(cc)-amount
sheet.cell(row=p+1, column=8).font=fontText1
sheet.cell(row=p+1, column=8).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=11).font=fontText1
sheet.cell(row=p+1, column=11).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=9).value=sheet.cell(row=p+1, column=7).value+sheet.cell(row=p+1, column=11).value
sheet.cell(row=p+1, column=9).font=fontText1
sheet.cell(row=p+1, column=9).number_format = '0.00_);(0.00)'
sheet.cell(row=p+1, column=10).value=sheet.cell(row=p+1, column=9).value-sheet.cell(row=p+1, column=8).value
sheet.cell(row=p+1, column=10).font=fontText1
sheet.cell(row=p+1, column=10).number_format = '0.00_ '
wb.save(filename5)

# In[ ]:

