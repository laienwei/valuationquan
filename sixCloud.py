#!/usr/bin/python3
# coding: utf-8

# In[ ]:
#在因为国证指数公司有新的指数发布后，399088和399378之前就有可能插入新的指数，本文件就需要更改，显然由于均值计算的缘故更改时需要用ctrl+H统一替换而绝非仅修改某一两个参数，切记切记！！！

import openpyxl

import publai1
pub=publai1.publai1()
with open(pub.filename,"r") as fp:
    lines=fp.readlines()
num=len(lines)
wb0 = openpyxl.load_workbook(r'./crawler/newworkbook.xlsx')
sheet0 = wb0['start_from_2025.2.28']
xx=sheet0.max_row-num #要读的数据从xx+1行开始
filedir3=['szseinnovation100indexmodel1.xlsx']
dirba1="./lai/valuationquan/szseinnovation100index"
filename3=dirba1+"/"+filedir3[0]
wb = openpyxl.load_workbook(filename3)
sheet = wb['myPEPB']
p=sheet.max_row
print(p)
cell = sheet['C3']
sum0=cell.value
for i in range(p-3):
    print(sum0)
    sum0=sum0+sheet.cell(row=i+4, column=3).value
print("over now")
print(sum0)
for i in range(num):
    print(xx+i+1)
    print(sheet0.cell(row=xx+i+1, column=24).value)
    #sum0=sum0+float(sheet0.cell(row=xx+i+1, column=24).value)
    sum0=sum0+sheet0.cell(row=xx+i+1, column=24).value
    sheet.cell(row=p+i+1, column=3).value=sheet0.cell(row=xx+i+1, column=24).value
    sheet.cell(row=p+i+1, column=1).value=p+i-1
    sheet.cell(row=p+i+1, column=2).value=lines[i]
    sheet.cell(row=p+i+1, column=4).value=sum0/(p+i-1)
wb.save(filename3)
wb.close()

filedir3=['CNIESG300indexmodel1.xlsx']
dirba1="./lai/valuationquan/CNIESG300index"
filename3=dirba1+"/"+filedir3[0]
wb = openpyxl.load_workbook(filename3)
sheet = wb['myPEPB']
p=sheet.max_row
print(p)
cell = sheet['C3']
sum0=cell.value
for i in range(p-3):
    print(sum0)
    sum0=sum0+sheet.cell(row=i+4, column=3).value
print("over now")
print(sum0)
for i in range(num):
    sum0=sum0+sheet0.cell(row=xx+i+1, column=134).value
    sheet.cell(row=p+i+1, column=3).value=sheet0.cell(row=xx+i+1, column=134).value
    sheet.cell(row=p+i+1, column=1).value=p+i-1
    sheet.cell(row=p+i+1, column=2).value=lines[i]
    sheet.cell(row=p+i+1, column=4).value=sum0/(p+i-1)
wb.save(filename3)
wb.close()
#####################################################################
filedir3=['szseinnovation100indexmodel1cn.xlsx']
dirba1="./lai/valuationquan/szseinnovation100index"
filename3=dirba1+"/"+filedir3[0]
wb = openpyxl.load_workbook(filename3)
sheet = wb['myPEPB']
p=sheet.max_row
print(p)
cell = sheet['C3']
sum0=cell.value
for i in range(p-3):
    print(sum0)
    sum0=sum0+sheet.cell(row=i+4, column=3).value
print("over now")
print(sum0)
for i in range(num):
    print(xx+i+1)
    print(sheet0.cell(row=xx+i+1, column=24).value)
    #sum0=sum0+float(sheet0.cell(row=xx+i+1, column=24).value)
    sum0=sum0+sheet0.cell(row=xx+i+1, column=24).value
    sheet.cell(row=p+i+1, column=3).value=sheet0.cell(row=xx+i+1, column=24).value
    sheet.cell(row=p+i+1, column=1).value=p+i-1
    sheet.cell(row=p+i+1, column=2).value=lines[i]
    sheet.cell(row=p+i+1, column=4).value=sum0/(p+i-1)
wb.save(filename3)
wb.close()

filedir3=['CNIESG300indexmodel1cn.xlsx']
dirba1="./lai/valuationquan/CNIESG300index"
filename3=dirba1+"/"+filedir3[0]
wb = openpyxl.load_workbook(filename3)
sheet = wb['myPEPB']
p=sheet.max_row
print(p)
cell = sheet['C3']
sum0=cell.value
for i in range(p-3):
    print(sum0)
    sum0=sum0+sheet.cell(row=i+4, column=3).value
print("over now")
print(sum0)
for i in range(num):
    sum0=sum0+sheet0.cell(row=xx+i+1, column=134).value
    sheet.cell(row=p+i+1, column=3).value=sheet0.cell(row=xx+i+1, column=134).value
    sheet.cell(row=p+i+1, column=1).value=p+i-1
    sheet.cell(row=p+i+1, column=2).value=lines[i]
    sheet.cell(row=p+i+1, column=4).value=sum0/(p+i-1)
wb.save(filename3)
wb.close()





